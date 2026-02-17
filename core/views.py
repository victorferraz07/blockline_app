from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q, Sum, Avg
from django.forms import modelformset_factory, inlineformset_factory
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.utils import timezone
import json

from .models import (
    ItemEstoque, ProdutoFabricado, Recebimento,
    DocumentoProdutoFabricado, Componente, Setor, Fornecedor,
    ImagemProdutoFabricado, ImagemItemEstoque, ItemFornecedor, Expedicao, ItemExpedido, DocumentoExpedicao, ImagemExpedicao,
    Empresa, PerfilUsuario,
    TaskQuantidadeFeita, TaskHistorico,
    JornadaTrabalho, RegistroPonto, ResumoMensal, AbonoDia,
    MovimentacaoEstoque, Cliente,
)
from .forms import (
    ItemEstoqueForm, RetiradaItemForm, AdicaoItemForm, RecebimentoForm,
    ProdutoFabricadoForm, DocumentoProdutoForm, ComponenteForm, ProducaoForm,
    ImagemProdutoForm, ItemFornecedorForm, ExpedicaoForm, ItemExpedidoForm, DocumentoExpedicaoForm, ImagemExpedicaoForm,
    ClienteForm, FornecedorForm
)
from .decorators import superuser_required, filter_by_empresa, get_user_empresa

def get_empresas_permitidas(user):
    if user.is_superuser:
        return Empresa.objects.all()
    try:
        if hasattr(user, 'perfil'):
            return user.perfil.empresas_permitidas.all()
    except PerfilUsuario.DoesNotExist:
        pass
    return Empresa.objects.none()

# core/views.py
@login_required
def dashboard(request):
    from django.db.models import Sum
    from datetime import datetime, timedelta

    # Estatísticas gerais
    total_itens_estoque = ItemEstoque.objects.count()
    total_produtos_fabricados = ProdutoFabricado.objects.count()
    total_recebimentos = Recebimento.objects.count()
    total_expedicoes = Expedicao.objects.count()

    # Quantidade total em estoque
    quantidade_total_estoque = ItemEstoque.objects.aggregate(total=Sum('quantidade'))['total'] or 0

    # Atividades recentes
    ultimos_recebimentos = Recebimento.objects.order_by('-data_recebimento')[:5]
    ultimas_expedicoes = Expedicao.objects.order_by('-data_expedicao')[:5]

    # Combinar e ordenar atividades por data
    atividades = []

    for recebimento in ultimos_recebimentos:
        atividades.append({
            'tipo': 'recebimento',
            'data': recebimento.data_recebimento,
            'objeto': recebimento
        })

    for expedicao in ultimas_expedicoes:
        atividades.append({
            'tipo': 'expedicao',
            'data': expedicao.data_expedicao,
            'objeto': expedicao
        })

    # Ordenar por data (mais recente primeiro)
    atividades.sort(key=lambda x: x['data'], reverse=True)
    atividades = atividades[:10]  # Limitar a 10 atividades

    contexto = {
        'total_itens_estoque': total_itens_estoque,
        'total_produtos_fabricados': total_produtos_fabricados,
        'total_recebimentos': total_recebimentos,
        'total_expedicoes': total_expedicoes,
        'quantidade_total_estoque': quantidade_total_estoque,
        'atividades': atividades,
    }
    return render(request, 'core/dashboard.html', contexto)

# --- Views de Estoque ---
@login_required
def lista_estoque(request):
    # SEGURANÇA: Filtrar apenas itens da empresa do usuário
    itens = ItemEstoque.objects.all()
    # Nota: ItemEstoque não tem campo empresa, então mostra todos
    # Se precisar filtro por empresa, adicione campo empresa ao modelo

    # Busca funcional
    query = request.GET.get('q', '').strip()
    if query:
        itens = itens.filter(
            Q(nome__icontains=query) |
            Q(descricao__icontains=query) |
            Q(tipo__icontains=query)
        )

    # Filtro por tipo
    tipo_filtro = request.GET.get('tipo', '').strip()
    if tipo_filtro:
        itens = itens.filter(tipo=tipo_filtro)

    # Ordenação
    ordenacao = request.GET.get('ordenar', 'nome')
    if ordenacao == 'nome':
        itens = itens.order_by('nome')
    elif ordenacao == 'estoque_asc':
        itens = itens.order_by('quantidade', 'nome')
    elif ordenacao == 'estoque_desc':
        itens = itens.order_by('-quantidade', 'nome')
    elif ordenacao == 'tipo':
        itens = itens.order_by('tipo', 'nome')
    elif ordenacao == 'data_desc':
        itens = itens.order_by('-id')
    else:
        itens = itens.order_by('nome')

    # Estatísticas
    total_itens = itens.count()
    itens_esgotados = itens.filter(quantidade=0).count()
    itens_saudaveis = itens.filter(quantidade__gte=10).count()

    contexto = {
        'itens': itens,
        'total_itens': total_itens,
        'itens_esgotados': itens_esgotados,
        'itens_saudaveis': itens_saudaveis,
        'query': query,
        'ordenacao': ordenacao,
        'tipo_filtro': tipo_filtro,
    }
    return render(request, 'core/lista_estoque.html', contexto)

@login_required
def adicionar_item(request):
    if request.method == 'POST':
        form = ItemEstoqueForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Novo item adicionado com sucesso!')
            return redirect('lista_estoque')
    else:
        form = ItemEstoqueForm()
    contexto = {'form': form, 'titulo': 'Adicionar Novo Item ao Estoque'}
    return render(request, 'core/adicionar_item.html', contexto)

@login_required
def gerenciar_item(request, pk):
    from datetime import timedelta

    item = get_object_or_404(ItemEstoque, pk=pk)
    FornecedorItemFormSet = inlineformset_factory(ItemEstoque, ItemFornecedor, form=ItemFornecedorForm, extra=1, can_delete=True)

    if request.method == 'POST':
        form = ItemEstoqueForm(request.POST, request.FILES, instance=item)
        formset_fornecedores = FornecedorItemFormSet(request.POST, instance=item, prefix='fornecedores')

        if form.is_valid() and formset_fornecedores.is_valid():
            form.save()
            formset_fornecedores.save()
            messages.success(request, 'Item atualizado com sucesso!')
            return redirect('gerenciar_item', pk=item.pk)
        else:
            # Exibir erros de validação
            if not formset_fornecedores.is_valid():
                for form_error in formset_fornecedores.errors:
                    if form_error:
                        messages.error(request, f'Erro no formulário de fornecedores: {form_error}')
                for non_form_error in formset_fornecedores.non_form_errors():
                    messages.error(request, f'Erro: {non_form_error}')
    else:
        form = ItemEstoqueForm(instance=item)
        formset_fornecedores = FornecedorItemFormSet(instance=item, prefix='fornecedores')

    # Estatísticas e dados adicionais
    # 1. Valor total em estoque
    ultimo_preco = item.itemfornecedor_set.order_by('-data_cotacao').first()
    valor_total_estoque = 0
    if ultimo_preco:
        valor_total_estoque = item.quantidade * float(ultimo_preco.valor_pago)

    # 2. Custo médio unitário
    custo_medio = item.itemfornecedor_set.aggregate(media=Avg('valor_pago'))['media'] or 0

    # 3. Movimentações nos últimos 30 dias
    data_limite = timezone.now() - timedelta(days=30)
    movimentacoes_recentes = item.movimentacoes.filter(data_hora__gte=data_limite)
    total_entradas = movimentacoes_recentes.filter(tipo='entrada').aggregate(total=Sum('quantidade'))['total'] or 0
    total_saidas = movimentacoes_recentes.filter(tipo='saida').aggregate(total=Sum('quantidade'))['total'] or 0

    # 4. Histórico completo
    historico_movimentacoes = item.movimentacoes.all()[:20]  # Últimas 20 movimentações

    # 5. Galeria de imagens
    galeria_imagens = item.imagens.all()

    # 6. Dados para o gráfico de evolução (últimos 30 dias)
    evolucao_estoque = []
    for i in range(29, -1, -1):
        data = timezone.now() - timedelta(days=i)
        data_inicio = data.replace(hour=0, minute=0, second=0, microsecond=0)
        data_fim = data_inicio + timedelta(days=1)

        # Calcular estoque naquela data
        movimentacoes_ate_data = item.movimentacoes.filter(data_hora__lt=data_fim)
        entradas = movimentacoes_ate_data.filter(tipo='entrada').aggregate(total=Sum('quantidade'))['total'] or 0
        saidas = movimentacoes_ate_data.filter(tipo='saida').aggregate(total=Sum('quantidade'))['total'] or 0
        estoque_dia = entradas - saidas

        evolucao_estoque.append({
            'data': data.strftime('%d/%m'),
            'quantidade': estoque_dia
        })

    contexto = {
        'form': form,
        'item': item,
        'formset_fornecedores': formset_fornecedores,
        'retirada_form': RetiradaItemForm(),
        'adicao_form': AdicaoItemForm(),
        # Estatísticas
        'valor_total_estoque': valor_total_estoque,
        'custo_medio': custo_medio,
        'total_entradas_30d': total_entradas,
        'total_saidas_30d': total_saidas,
        # Histórico e galeria
        'historico_movimentacoes': historico_movimentacoes,
        'galeria_imagens': galeria_imagens,
        # Gráfico (convertido para JSON)
        'evolucao_estoque': json.dumps(evolucao_estoque),
    }
    return render(request, 'core/gerenciar_item.html', contexto)

@login_required
def retirar_item(request, pk):
    item = get_object_or_404(ItemEstoque, pk=pk)
    if request.method == 'POST':
        form = RetiradaItemForm(request.POST)
        if form.is_valid():
            quantidade_a_retirar = form.cleaned_data['quantidade']
            observacoes = form.cleaned_data.get('observacoes', '')
            if quantidade_a_retirar > item.quantidade:
                messages.error(request, f'Não é possível retirar {quantidade_a_retirar}. Quantidade em estoque: {item.quantidade}.')
            else:
                item.quantidade -= quantidade_a_retirar
                item.save()

                # Registrar movimentação
                MovimentacaoEstoque.objects.create(
                    item=item,
                    tipo='saida',
                    quantidade=quantidade_a_retirar,
                    usuario=request.user,
                    observacoes=observacoes
                )

                messages.success(request, f'{quantidade_a_retirar} unidade(s) de {item.nome} retiradas com sucesso.')
    return redirect('gerenciar_item', pk=item.pk)

@login_required
def adicionar_estoque(request, pk):
    item = get_object_or_404(ItemEstoque, pk=pk)
    if request.method == 'POST':
        form = AdicaoItemForm(request.POST)
        if form.is_valid():
            quantidade_a_adicionar = form.cleaned_data['quantidade']
            observacoes = form.cleaned_data.get('observacoes', '')
            item.quantidade += quantidade_a_adicionar
            item.save()

            # Registrar movimentação
            MovimentacaoEstoque.objects.create(
                item=item,
                tipo='entrada',
                quantidade=quantidade_a_adicionar,
                usuario=request.user,
                observacoes=observacoes
            )

            messages.success(request, f'{quantidade_a_adicionar} unidade(s) de {item.nome} adicionadas com sucesso.')
    return redirect('gerenciar_item', pk=item.pk)

@login_required
@permission_required('core.delete_itemestoque', raise_exception=True)
def excluir_item(request, pk):
    item = get_object_or_404(ItemEstoque, pk=pk)
    if request.method == 'POST':
        nome_item = item.nome
        item.delete()
        messages.success(request, f'Item "{nome_item}" foi excluído com sucesso.')
        return redirect('lista_estoque')
    return redirect('gerenciar_item', pk=pk)

@login_required
def duplicar_item(request, pk):
    from django.core.files.base import ContentFile
    import os

    item_original = get_object_or_404(ItemEstoque, pk=pk)

    # Gerar nome único para a cópia
    nome_base = f"{item_original.nome} (Cópia)"
    nome_final = nome_base
    contador = 1

    while ItemEstoque.objects.filter(nome=nome_final).exists():
        nome_final = f"{nome_base} {contador}"
        contador += 1

    # Criar cópia do item
    item_duplicado = ItemEstoque.objects.create(
        tipo=item_original.tipo,
        nome=nome_final,
        descricao=item_original.descricao,
        quantidade=0,  # Começa com quantidade zerada
        local_armazenamento=item_original.local_armazenamento,
        is_produto_fabricado=item_original.is_produto_fabricado
    )

    # Copiar foto principal se existir
    if item_original.foto_principal and os.path.exists(item_original.foto_principal.path):
        with open(item_original.foto_principal.path, 'rb') as f:
            item_duplicado.foto_principal.save(
                os.path.basename(item_original.foto_principal.name),
                ContentFile(f.read()),
                save=True
            )

    # Copiar documentação se existir
    if item_original.documentacao and os.path.exists(item_original.documentacao.path):
        with open(item_original.documentacao.path, 'rb') as f:
            item_duplicado.documentacao.save(
                os.path.basename(item_original.documentacao.name),
                ContentFile(f.read()),
                save=True
            )

    # Copiar fornecedores associados
    for item_fornecedor in item_original.itemfornecedor_set.all():
        ItemFornecedor.objects.create(
            item_estoque=item_duplicado,
            fornecedor=item_fornecedor.fornecedor,
            fornecedor_nome=item_fornecedor.fornecedor_nome,
            valor_pago=item_fornecedor.valor_pago,
            data_cotacao=item_fornecedor.data_cotacao
        )

    messages.success(request, f'Item "{item_duplicado.nome}" foi criado com sucesso!')
    return redirect('gerenciar_item', pk=item_duplicado.pk)


# --- Views de Recebimento ---

@login_required
def lista_recebimentos(request):
    from datetime import timedelta
    from decimal import Decimal

    query = request.GET.get('q', '')
    status_filtro = request.GET.get('status', '')
    ordenacao = request.GET.get('ordenar', 'data_desc')

    # Filtrar recebimentos
    recebimentos = Recebimento.objects.all()

    # Aplicar filtro de busca
    if query:
        recebimentos = recebimentos.filter(
            Q(numero_nota_fiscal__icontains=query) |
            Q(fornecedor__nome__icontains=query) |
            Q(usuario__username__icontains=query) |
            Q(setor__nome__icontains=query)
        )

    # Aplicar filtro de status
    if status_filtro:
        recebimentos = recebimentos.filter(status=status_filtro)

    # Aplicar ordenação
    if ordenacao == 'data_desc':
        recebimentos = recebimentos.order_by('-data_recebimento')
    elif ordenacao == 'data_asc':
        recebimentos = recebimentos.order_by('data_recebimento')
    elif ordenacao == 'fornecedor':
        recebimentos = recebimentos.order_by('fornecedor__nome')
    elif ordenacao == 'valor_desc':
        recebimentos = recebimentos.order_by('-valor_total')
    elif ordenacao == 'status':
        recebimentos = recebimentos.order_by('status')

    # Estatísticas
    total_recebimentos = Recebimento.objects.count()

    # Valor total do mês atual
    data_limite = timezone.now() - timedelta(days=30)
    recebimentos_mes = Recebimento.objects.filter(data_recebimento__gte=data_limite)
    valor_total_mes = recebimentos_mes.aggregate(total=Sum('valor_total'))['total'] or Decimal('0.00')

    # Recebimentos aguardando
    aguardando_conferencia = Recebimento.objects.filter(status='aguardando').count()

    contexto = {
        'recebimentos': recebimentos,
        'query': query,
        'status_filtro': status_filtro,
        'ordenacao': ordenacao,
        'total_recebimentos': total_recebimentos,
        'valor_total_mes': valor_total_mes,
        'aguardando_conferencia': aguardando_conferencia,
    }
    return render(request, 'core/lista_recebimentos.html', contexto)

@login_required
# @permission_required('core.add_recebimento', raise_exception=True)
def registrar_recebimento(request):
    if request.method == 'POST':
        form = RecebimentoForm(request.POST, request.FILES)
        # Remover campo empresa do formulário (será atribuído automaticamente)
        if 'empresa' in form.fields:
            del form.fields['empresa']

        if form.is_valid():
            recebimento = form.save(commit=False)

            # Atribuir primeira empresa disponível
            primeira_empresa = Empresa.objects.first()
            if primeira_empresa:
                recebimento.empresa = primeira_empresa
            else:
                messages.error(request, "Nenhuma empresa cadastrada no sistema.")
                return redirect('dashboard')

            recebimento.usuario = request.user
            recebimento.save()
            messages.success(request, 'Recebimento registrado com sucesso!')
            return redirect('lista_recebimentos')
        else:
            # Mostrar erros de validação com nomes de campos traduzidos
            for field, errors in form.errors.items():
                if field in form.fields:
                    field_label = form.fields[field].label
                else:
                    field_label = field
                for error in errors:
                    messages.error(request, f"Campo '{field_label}': {error}")
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, f"Erro: {error}")
    else:
        form = RecebimentoForm()
        # Remover campo empresa do formulário (será atribuído automaticamente)
        if 'empresa' in form.fields:
            del form.fields['empresa']

    contexto = {
        'form': form,
        'titulo': 'Registrar Novo Recebimento',
        'ultimos_recebimentos': Recebimento.objects.all().order_by('-data_recebimento')[:5]
    }
    return render(request, 'core/registrar_recebimento.html', contexto)

@login_required
def detalhe_recebimento(request, pk):
    empresas_permitidas = get_empresas_permitidas(request.user)
    recebimento = get_object_or_404(Recebimento, pk=pk)
    contexto = {'recebimento': recebimento}

    return render(request, 'core/detalhe_recebimento.html', contexto)

@login_required
def editar_recebimento(request, pk):
    recebimento = get_object_or_404(Recebimento, pk=pk)
    if request.method == 'POST':
        post_data = request.POST.copy()
        fornecedor_val = post_data.get('fornecedor')
        if fornecedor_val and not fornecedor_val.isnumeric():
            novo_fornecedor, created = Fornecedor.objects.get_or_create(nome=fornecedor_val)
            post_data['fornecedor'] = novo_fornecedor.pk
        form = RecebimentoForm(post_data, request.FILES, instance=recebimento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Recebimento atualizado com sucesso!')
            return redirect('lista_recebimentos')
    else:
        form = RecebimentoForm(instance=recebimento)
        
    contexto = {
        'form': form,
        'titulo': f'Editar Recebimento (NF: {recebimento.numero_nota_fiscal})',
        'is_editing': True,
        'recebimento': recebimento # Garante que o objeto está disponível
    }
    return render(request, 'core/registrar_recebimento.html', contexto)

# --- Views de Produtos Fabricados ---

@login_required
def lista_produtos(request):
    produtos = ProdutoFabricado.objects.select_related('item_associado').all()

    # Busca
    query = request.GET.get('q', '').strip()
    if query:
        produtos = produtos.filter(
            Q(nome__icontains=query) |
            Q(descricao__icontains=query)
        )

    # Ordenação
    ordenacao = request.GET.get('ordenar', 'nome')
    if ordenacao == 'nome':
        produtos = produtos.order_by('nome')
    elif ordenacao == 'data_desc':
        produtos = produtos.order_by('-id')  # Mais recentes primeiro
    elif ordenacao == 'data_asc':
        produtos = produtos.order_by('id')   # Mais antigos primeiro
    else:
        produtos = produtos.order_by('nome')  # Fallback

    contexto = {
        'produtos': produtos,
        'query': query,
        'ordenacao': ordenacao,
    }
    return render(request, 'core/lista_produtos.html', contexto)

@login_required
def detalhe_produto(request, pk):
    from datetime import datetime
    from django.db.models import Sum, Avg

    produto = get_object_or_404(ProdutoFabricado, pk=pk)
    componentes = produto.componente_set.all()

    # Lógica do Simulador
    form = ProducaoForm(request.GET or None)
    qtd_a_produzir = 1 # Valor padrão
    if form.is_valid():
        qtd_a_produzir = form.cleaned_data['quantidade_a_produzir']

    # Adicionando os cálculos a cada componente
    custo_total_estimado = 0
    todos_tem_estoque = True

    for componente in componentes:
        componente.total_necessario = componente.quantidade_necessaria * qtd_a_produzir
        componente.estoque_suficiente = componente.item_estoque.quantidade >= componente.total_necessario

        if not componente.estoque_suficiente:
            todos_tem_estoque = False

        # Calcular custo estimado usando o último preço do fornecedor
        ultimo_preco = componente.item_estoque.itemfornecedor_set.order_by('-data_cotacao').first()
        if ultimo_preco:
            componente.custo_unitario = ultimo_preco.valor_pago
            componente.custo_total = componente.custo_unitario * componente.total_necessario
            custo_total_estimado += componente.custo_total
        else:
            componente.custo_unitario = 0
            componente.custo_total = 0

    # Estatísticas do produto
    estoque_atual = produto.item_associado.quantidade
    data_criacao = produto.item_associado.data_criacao
    dias_desde_criacao = (datetime.now().date() - data_criacao.date()).days if data_criacao else 0

    # Média mensal aproximada (baseada no estoque atual e dias desde criação)
    if dias_desde_criacao > 0:
        media_mensal = (estoque_atual / dias_desde_criacao) * 30
    else:
        media_mensal = 0

    contexto = {
        'produto': produto,
        'componentes': componentes,
        'producao_form': form,
        'qtd_a_produzir': qtd_a_produzir,
        'custo_total_estimado': round(custo_total_estimado, 2),
        'custo_unitario_estimado': round(custo_total_estimado / qtd_a_produzir, 2) if qtd_a_produzir > 0 else 0,
        'todos_tem_estoque': todos_tem_estoque,
        # Estatísticas
        'estoque_atual': estoque_atual,
        'data_criacao': data_criacao,
        'media_mensal': round(media_mensal, 1),
    }
    return render(request, 'core/detalhe_produto.html', contexto)

@login_required
def adicionar_produto(request):
    DocumentoFormSet = modelformset_factory(DocumentoProdutoFabricado, form=DocumentoProdutoForm, extra=1, can_delete=False)
    
    if request.method == 'POST':
        form = ProdutoFabricadoForm(request.POST, request.FILES)
        formset = DocumentoFormSet(request.POST, request.FILES, queryset=DocumentoProdutoFabricado.objects.none())

        if form.is_valid() and formset.is_valid():
            produto_base = form.save(commit=False)
            
            # Criamos o item de estoque associado
            item_estoque_associado = ItemEstoque.objects.create(
                nome=produto_base.nome,
                descricao=f"Produto acabado: {produto_base.descricao}",
                tipo='produto_acabado',
                is_produto_fabricado=True,
                # A LINHA DA CORREÇÃO ESTÁ AQUI:
                # Copiamos a foto do formulário do produto para o novo item de estoque.
                foto_principal=produto_base.foto_principal 
            )
            
            # Agora, linkamos os dois e salvamos o produto
            produto_base.item_associado = item_estoque_associado
            produto_base.save()
            
            # ... (resto da lógica para salvar os documentos)
            for doc_form in formset:
                if doc_form.cleaned_data and doc_form.cleaned_data.get('documento'):
                    documento = doc_form.save(commit=False)
                    documento.produto = produto_base
                    documento.save()

            messages.success(request, f'Produto "{produto_base.nome}" e seus documentos foram criados com sucesso!')
            return redirect('lista_produtos')
    else:
        form = ProdutoFabricadoForm()
        formset = DocumentoFormSet(queryset=DocumentoProdutoFabricado.objects.none())
        
    contexto = {'form': form, 'formset': formset, 'titulo': 'Criar Novo Produto'}
    return render(request, 'core/form_produto.html', contexto)

@login_required
def editar_produto(request, pk):
    produto = get_object_or_404(ProdutoFabricado, pk=pk)
    
    ComponenteFormSet = inlineformset_factory(ProdutoFabricado, Componente, form=ComponenteForm, extra=0, can_delete=True)
    DocumentoFormSet = inlineformset_factory(ProdutoFabricado, DocumentoProdutoFabricado, form=DocumentoProdutoForm, extra=0, can_delete=True)
    ImagemFormSet = inlineformset_factory(ProdutoFabricado, ImagemProdutoFabricado, form=ImagemProdutoForm, extra=1, can_delete=True)

    if request.method == 'POST':
        form = ProdutoFabricadoForm(request.POST, request.FILES, instance=produto)
        componente_formset = ComponenteFormSet(request.POST, instance=produto, prefix='componentes')
        documento_formset = DocumentoFormSet(request.POST, request.FILES, instance=produto, prefix='documentos')
        imagem_formset = ImagemFormSet(request.POST, request.FILES, instance=produto, prefix='imagens')

        if form.is_valid() and componente_formset.is_valid() and documento_formset.is_valid() and imagem_formset.is_valid():
            form.save()
            componente_formset.save()
            documento_formset.save()
            imagem_formset.save()
            
            messages.success(request, f'Produto "{produto.nome}" atualizado com sucesso!')
            return redirect('detalhe_produto', pk=produto.pk)
    else:
        form = ProdutoFabricadoForm(instance=produto)
        componente_formset = ComponenteFormSet(instance=produto, prefix='componentes')
        documento_formset = DocumentoFormSet(instance=produto, prefix='documentos')
        imagem_formset = ImagemFormSet(instance=produto, prefix='imagens')

    contexto = {
        'form': form,
        'componente_formset': componente_formset,
        'documento_formset': documento_formset,
        'imagem_formset': imagem_formset,
        'produto': produto,
    }
    return render(request, 'core/editar_produto.html', contexto)

@login_required
@transaction.atomic
def finalizar_producao(request, pk):
    produto = get_object_or_404(ProdutoFabricado, pk=pk)
    if request.method == 'POST':
        form = ProducaoForm(request.POST)
        if form.is_valid():
            qtd_a_produzir = form.cleaned_data['quantidade_a_produzir']
            componentes_insuficientes = []
            for componente in produto.componente_set.all():
                total_necessario = componente.quantidade_necessaria * qtd_a_produzir
                if componente.item_estoque.quantidade < total_necessario:
                    componentes_insuficientes.append(componente.item_estoque.nome)
            if componentes_insuficientes:
                msg_erro = f'Produção não pode ser iniciada. Estoque insuficiente para: {", ".join(componentes_insuficientes)}.'
                messages.error(request, msg_erro)
                return redirect('detalhe_produto', pk=pk)
            for componente in produto.componente_set.all():
                componente.item_estoque.quantidade -= componente.quantidade_necessaria * qtd_a_produzir
                componente.item_estoque.save()
            if hasattr(produto, 'item_associado') and produto.item_associado:
                produto.item_associado.quantidade += qtd_a_produzir
                produto.item_associado.save()
            else:
                messages.error(request, f'Erro crítico: O produto "{produto.nome}" não está associado a um item de estoque.')
                return redirect('detalhe_produto', pk=pk)
            messages.success(request, f'{qtd_a_produzir} unidade(s) de "{produto.nome}" foram produzidas e adicionadas ao estoque!')
            return redirect('detalhe_produto', pk=pk)
    return redirect('detalhe_produto', pk=pk)

# --- Views da Expedição ---

@login_required
def lista_expedicoes(request):
    from django.db.models import Sum, Count
    from datetime import datetime, timedelta

    # Pega todas as expedições da empresa do usuário
    expedicoes = Expedicao.objects.all()

    # Filtro por pesquisa (cliente ou nota fiscal)
    search_query = request.GET.get('search', '')
    if search_query:
        expedicoes = expedicoes.filter(
            Q(cliente__icontains=search_query) |
            Q(nota_fiscal__icontains=search_query)
        )

    # Ordenação
    order_by = request.GET.get('order_by', '-data_expedicao')
    valid_order_fields = ['-data_expedicao', 'data_expedicao', 'cliente', '-cliente', 'nota_fiscal', '-nota_fiscal']
    if order_by in valid_order_fields:
        expedicoes = expedicoes.order_by(order_by)

    # Estatísticas
    total_expedicoes = Expedicao.objects.count()

    # Total de itens expedidos (soma de todas as quantidades)
    total_itens_expedidos = ItemExpedido.objects.aggregate(
        total=Sum('quantidade')
    )['total'] or 0

    # Expedições dos últimos 30 dias
    data_limite = datetime.now() - timedelta(days=30)
    expedicoes_30_dias = Expedicao.objects.filter(data_expedicao__gte=data_limite).count()

    contexto = {
        'expedicoes': expedicoes,
        'total_expedicoes': total_expedicoes,
        'total_itens_expedidos': total_itens_expedidos,
        'expedicoes_30_dias': expedicoes_30_dias,
        'search_query': search_query,
        'order_by': order_by,
    }
    return render(request, 'core/lista_expedicoes.html', contexto)

@login_required
def detalhe_expedicao(request, pk):
    expedicao = get_object_or_404(Expedicao, pk=pk)
    return render(request, 'core/detalhe_expedicao.html', {'expedicao': expedicao})

@login_required
@transaction.atomic
def registrar_expedicao(request):
    ItemExpedidoFormSet = inlineformset_factory(Expedicao, ItemExpedido, form=ItemExpedidoForm, extra=1, can_delete=False)
    DocumentoExpedicaoFormSet = inlineformset_factory(Expedicao, DocumentoExpedicao, form=DocumentoExpedicaoForm, extra=1, can_delete=False)
    ImagemExpedicaoFormSet = inlineformset_factory(Expedicao, ImagemExpedicao, form=ImagemExpedicaoForm, extra=1, can_delete=False)

    if request.method == 'POST':
        form = ExpedicaoForm(request.POST, request.FILES)
        # Remover campo empresa do formulário (será atribuído automaticamente)
        if 'empresa' in form.fields:
            del form.fields['empresa']

        item_formset = ItemExpedidoFormSet(request.POST, prefix='itens')
        documento_formset = DocumentoExpedicaoFormSet(request.POST, request.FILES, prefix='documentos')
        imagem_formset = ImagemExpedicaoFormSet(request.POST, request.FILES, prefix='imagens')

        if form.is_valid() and item_formset.is_valid() and documento_formset.is_valid() and imagem_formset.is_valid():

            # --- Verificação de Estoque ---
            pode_prosseguir = True
            for form_item in item_formset:
                if form_item.cleaned_data and form_item.cleaned_data.get('produto'):
                    produto = form_item.cleaned_data['produto']
                    quantidade = form_item.cleaned_data['quantidade']
                    if quantidade > produto.item_associado.quantidade:
                        messages.error(request, f"Estoque insuficiente para {produto.nome}. Disponível: {produto.item_associado.quantidade}")
                        pode_prosseguir = False

            if not pode_prosseguir:
                # Se não puder prosseguir, renderiza a página de novo com os erros
                contexto = {'form': form, 'item_formset': item_formset, 'documento_formset': documento_formset, 'imagem_formset': imagem_formset}
                return render(request, 'core/registrar_expedicao.html', contexto)

            # --- Se tudo estiver OK, salva e dá baixa no estoque ---
            expedicao = form.save(commit=False)

            # Atribuir primeira empresa disponível
            primeira_empresa = Empresa.objects.first()
            if primeira_empresa:
                expedicao.empresa = primeira_empresa
            else:
                messages.error(request, "Nenhuma empresa cadastrada no sistema.")
                return redirect('dashboard')

            expedicao.usuario = request.user
            expedicao.save()

            item_formset.instance = expedicao
            item_formset.save()

            documento_formset.instance = expedicao
            documento_formset.save()
            
            imagem_formset.instance = expedicao
            imagem_formset.save()
            
            # Baixa no estoque
            for form_item in item_formset:
                if form_item.cleaned_data and form_item.cleaned_data.get('produto'):
                    produto = form_item.cleaned_data['produto']
                    quantidade = form_item.cleaned_data['quantidade']
                    produto.item_associado.quantidade -= quantidade
                    produto.item_associado.save()

            messages.success(request, f"Expedição #{expedicao.pk} registrada com sucesso!")
            return redirect('lista_expedicoes')
    else:
        form = ExpedicaoForm()
        item_formset = ItemExpedidoFormSet(prefix='itens')
        documento_formset = DocumentoExpedicaoFormSet(prefix='documentos')
        imagem_formset = ImagemExpedicaoFormSet(prefix='imagens')

    contexto = {
        'form': form,
        'item_formset': item_formset,
        'documento_formset': documento_formset,
        'imagem_formset': imagem_formset,
        'titulo': 'Registrar Nova Expedição'
    }
    return render(request, 'core/registrar_expedicao.html', contexto)

@login_required
def editar_expedicao(request, pk):
    expedicao = get_object_or_404(Expedicao, pk=pk)
    
    ItemExpedidoFormSet = inlineformset_factory(Expedicao, ItemExpedido, form=ItemExpedidoForm, extra=1, can_delete=True)
    DocumentoExpedicaoFormSet = inlineformset_factory(Expedicao, DocumentoExpedicao, form=DocumentoExpedicaoForm, extra=1, can_delete=True)
    ImagemExpedicaoFormSet = inlineformset_factory(Expedicao, ImagemExpedicao, form=ImagemExpedicaoForm, extra=1, can_delete=True)

    if request.method == 'POST':
        form = ExpedicaoForm(request.POST, instance=expedicao)
        item_formset = ItemExpedidoFormSet(request.POST, instance=expedicao, prefix='itens')
        documento_formset = DocumentoExpedicaoFormSet(request.POST, request.FILES, instance=expedicao, prefix='documentos')
        imagem_formset = ImagemExpedicaoFormSet(request.POST, request.FILES, instance=expedicao, prefix='imagens')

        if form.is_valid() and item_formset.is_valid() and documento_formset.is_valid() and imagem_formset.is_valid():
            form.save()
            item_formset.save()
            documento_formset.save()
            imagem_formset.save()
            
            messages.success(request, f"Expedição #{expedicao.pk} atualizada com sucesso!")
            return redirect('detalhe_expedicao', pk=expedicao.pk)
    else:
        form = ExpedicaoForm(instance=expedicao)
        item_formset = ItemExpedidoFormSet(instance=expedicao, prefix='itens')
        documento_formset = DocumentoExpedicaoFormSet(instance=expedicao, prefix='documentos')
        imagem_formset = ImagemExpedicaoFormSet(instance=expedicao, prefix='imagens')

    contexto = {
        'titulo': f'Editando Expedição #{expedicao.pk}',
        'expedicao': expedicao,
        'form': form,
        'item_formset': item_formset,
        'documento_formset': documento_formset,
        'imagem_formset': imagem_formset
    }
    return render(request, 'core/editar_expedicao.html', contexto)

@login_required
def excluir_recebimento(request, pk):
    recebimento = get_object_or_404(Recebimento, pk=pk)
    if request.method == 'POST':
        recebimento.delete()
        messages.success(request, f"Registro de recebimento #{pk} foi excluído com sucesso.")
        return redirect('lista_recebimentos')
    # Se não for POST, apenas redireciona de volta para os detalhes
    return redirect('detalhe_recebimento', pk=pk)

@login_required
def excluir_produto(request, pk):
    produto = get_object_or_404(ProdutoFabricado, pk=pk)
    if request.method == 'POST':
        # Antes de deletar o produto, deletamos o item de estoque associado
        if produto.item_associado:
            produto.item_associado.delete()

        nome_produto = produto.nome
        produto.delete()
        messages.success(request, f'Produto "{nome_produto}" foi excluído com sucesso.')
        return redirect('lista_produtos')

    return redirect('detalhe_produto', pk=pk)

@login_required
def excluir_expedicao(request, pk):
    expedicao = get_object_or_404(Expedicao, pk=pk)
    if request.method == 'POST':
        expedicao.delete()
        messages.success(request, f"Expedição #{pk} foi excluída com sucesso.")
        return redirect('lista_expedicoes')
    
    return redirect('detalhe_expedicao', pk=pk)


# ==================================================
# VIEWS - SISTEMA DE PLANEJAMENTO DE PROJETOS
# ==================================================

@login_required
def roadmap_timeline(request):
    """Visualização principal: Timeline/Roadmap estilo Gantt"""
    from .models import Project, Milestone, Sprint, ProjectTask
    from django.contrib.auth.models import User

    # Filtros
    project_id = request.GET.get('project')
    view_mode = request.GET.get('view', 'month')  # 'month' ou 'week'

    # Buscar projetos — filtra por membros (superusuário vê todos)
    if request.user.is_superuser:
        projects = Project.objects.all().order_by('ordem')
    else:
        projects = Project.objects.filter(membros=request.user).order_by('ordem')

    project_selecionado = None

    if project_id:
        if request.user.is_superuser:
            project_selecionado = get_object_or_404(Project, id=project_id)
        else:
            project_selecionado = get_object_or_404(Project, id=project_id, membros=request.user)

    # Buscar tarefas
    tasks = ProjectTask.objects.filter(project=project_selecionado).select_related('milestone', 'project').prefetch_related('responsaveis') if project_selecionado else ProjectTask.objects.none()

    # Buscar milestones e sprints
    milestones = Milestone.objects.filter(project=project_selecionado) if project_selecionado else []
    sprints = Sprint.objects.filter(project=project_selecionado) if project_selecionado else []

    # Buscar usuários ativos
    users = User.objects.filter(is_active=True).order_by('first_name', 'username')

    # Organizar tarefas por milestone
    tasks_por_milestone = {}
    for milestone in milestones:
        tasks_por_milestone[milestone] = tasks.filter(milestone=milestone)

    # Tarefas sem milestone
    tasks_por_milestone[None] = tasks.filter(milestone__isnull=True)

    context = {
        'projects': projects,
        'project_selecionado': project_selecionado,
        'tasks': tasks,
        'tasks_por_milestone': tasks_por_milestone,
        'users': users,
        'milestones': milestones,
        'sprints': sprints,
        'view_mode': view_mode,
    }

    return render(request, 'core/roadmap_timeline.html', context)


# --- CRUD: Projects ---

@login_required
def criar_project(request):
    """Criar novo projeto"""
    from .models import Project
    from django.contrib.auth.models import User as UserModel

    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao', '')
        cor = request.POST.get('cor', '#6366f1')

        project = Project.objects.create(
            nome=nome,
            descricao=descricao,
            cor=cor,
            criado_por=request.user,
        )

        # Criador sempre é membro
        project.membros.add(request.user)

        # Adicionar membros selecionados
        membros_ids = request.POST.getlist('membros[]')
        if membros_ids:
            project.membros.add(*UserModel.objects.filter(id__in=membros_ids, is_active=True))

        messages.success(request, f'Projeto "{project.nome}" criado com sucesso!')
        return redirect('roadmap_timeline')

    users = UserModel.objects.filter(is_active=True).order_by('first_name', 'username')
    return render(request, 'core/form_project.html', {'users': users, 'membros_atuais': []})


@login_required
def editar_project(request, project_id):
    """Editar projeto existente"""
    from .models import Project
    from django.contrib.auth.models import User as UserModel

    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        project.nome = request.POST.get('nome')
        project.descricao = request.POST.get('descricao', '')
        project.cor = request.POST.get('cor', '#6366f1')
        project.save()

        # Somente o criador ou superusuário pode alterar membros
        if request.user == project.criado_por or request.user.is_superuser:
            membros_ids = request.POST.getlist('membros[]')
            novos_membros = UserModel.objects.filter(id__in=membros_ids, is_active=True)
            project.membros.set(novos_membros)
            # Criador sempre permanece membro
            if project.criado_por:
                project.membros.add(project.criado_por)

        messages.success(request, f'Projeto "{project.nome}" atualizado com sucesso!')
        return redirect('roadmap_timeline')

    users = UserModel.objects.filter(is_active=True).order_by('first_name', 'username')
    membros_atuais = list(project.membros.values_list('id', flat=True))
    eh_criador = request.user == project.criado_por or request.user.is_superuser
    context = {
        'project': project,
        'users': users,
        'membros_atuais': membros_atuais,
        'eh_criador': eh_criador,
    }
    return render(request, 'core/form_project.html', context)


@login_required
def excluir_project(request, project_id):
    """Excluir projeto"""
    from .models import Project

    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        nome = project.nome
        project.delete()
        messages.success(request, f'Projeto "{nome}" excluído com sucesso!')
        return redirect('roadmap_timeline')

    return redirect('roadmap_timeline')


# --- CRUD: Milestones ---

@login_required
def criar_milestone(request, project_id):
    """Criar novo milestone"""
    from .models import Project, Milestone

    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao', '')
        data_inicio = request.POST.get('data_inicio') or None
        data_fim = request.POST.get('data_fim') or None
        cor = request.POST.get('cor', '#10b981')
        status = request.POST.get('status', 'planejado')

        milestone = Milestone.objects.create(
            project=project,
            nome=nome,
            descricao=descricao,
            data_inicio=data_inicio,
            data_fim=data_fim,
            cor=cor,
            status=status,
        )

        messages.success(request, f'Milestone "{milestone.nome}" criado com sucesso!')
        return redirect('roadmap_timeline') + f'?project={project_id}'

    context = {'project': project}
    return render(request, 'core/form_milestone.html', context)


@login_required
def editar_milestone(request, milestone_id):
    """Editar milestone existente"""
    from .models import Milestone

    milestone = get_object_or_404(Milestone, id=milestone_id)

    if request.method == 'POST':
        milestone.nome = request.POST.get('nome')
        milestone.descricao = request.POST.get('descricao', '')
        milestone.data_inicio = request.POST.get('data_inicio') or None
        milestone.data_fim = request.POST.get('data_fim') or None
        milestone.cor = request.POST.get('cor', '#10b981')
        milestone.status = request.POST.get('status', 'planejado')
        milestone.save()

        messages.success(request, f'Milestone "{milestone.nome}" atualizado com sucesso!')
        return redirect('roadmap_timeline') + f'?project={milestone.project.id}'

    context = {'milestone': milestone}
    return render(request, 'core/form_milestone.html', context)


@login_required
def excluir_milestone(request, milestone_id):
    """Excluir milestone"""
    from .models import Milestone

    milestone = get_object_or_404(Milestone, id=milestone_id)
    project_id = milestone.project.id

    if request.method == 'POST':
        nome = milestone.nome
        milestone.delete()
        messages.success(request, f'Milestone "{nome}" excluído com sucesso!')
        return redirect('roadmap_timeline') + f'?project={project_id}'

    return redirect('roadmap_timeline')


# --- CRUD: Sprints ---

@login_required
def criar_sprint(request, project_id):
    """Criar novo sprint"""
    from .models import Project, Sprint

    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        objetivo = request.POST.get('objetivo', '')
        ativo = request.POST.get('ativo') == 'on'

        # Se ativo, desativar outros sprints do projeto
        if ativo:
            Sprint.objects.filter(project=project).update(ativo=False)

        sprint = Sprint.objects.create(
            project=project,
            nome=nome,
            data_inicio=data_inicio,
            data_fim=data_fim,
            objetivo=objetivo,
            ativo=ativo,
        )

        messages.success(request, f'Sprint "{sprint.nome}" criado com sucesso!')
        return redirect('roadmap_timeline') + f'?project={project_id}'

    context = {'project': project}
    return render(request, 'core/form_sprint.html', context)


@login_required
def editar_sprint(request, sprint_id):
    """Editar sprint existente"""
    from .models import Sprint

    sprint = get_object_or_404(Sprint, id=sprint_id)

    if request.method == 'POST':
        sprint.nome = request.POST.get('nome')
        sprint.data_inicio = request.POST.get('data_inicio')
        sprint.data_fim = request.POST.get('data_fim')
        sprint.objetivo = request.POST.get('objetivo', '')
        ativo = request.POST.get('ativo') == 'on'

        # Se ativo, desativar outros sprints do projeto
        if ativo:
            Sprint.objects.filter(project=sprint.project).exclude(id=sprint.id).update(ativo=False)

        sprint.ativo = ativo
        sprint.save()

        messages.success(request, f'Sprint "{sprint.nome}" atualizado com sucesso!')
        return redirect('roadmap_timeline') + f'?project={sprint.project.id}'

    context = {'sprint': sprint}
    return render(request, 'core/form_sprint.html', context)


@login_required
def excluir_sprint(request, sprint_id):
    """Excluir sprint"""
    from .models import Sprint

    sprint = get_object_or_404(Sprint, id=sprint_id)
    project_id = sprint.project.id

    if request.method == 'POST':
        nome = sprint.nome
        sprint.delete()
        messages.success(request, f'Sprint "{nome}" excluído com sucesso!')
        return redirect('roadmap_timeline') + f'?project={project_id}'

    return redirect('roadmap_timeline')


@login_required
def ativar_sprint(request, sprint_id):
    """Ativar sprint (desativa outros do projeto)"""
    from .models import Sprint

    sprint = get_object_or_404(Sprint, id=sprint_id)

    if request.method == 'POST':
        # Desativar todos os sprints do projeto
        Sprint.objects.filter(project=sprint.project).update(ativo=False)

        # Ativar este sprint
        sprint.ativo = True
        sprint.save()

        messages.success(request, f'Sprint "{sprint.nome}" ativado com sucesso!')

    return redirect('roadmap_timeline') + f'?project={sprint.project.id}'


# --- CRUD: Labels ---

@login_required
def criar_label(request, project_id):
    """Criar nova label"""
    from .models import Project, Label

    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        cor = request.POST.get('cor', '#6b7280')
        descricao = request.POST.get('descricao', '')

        label = Label.objects.create(
            project=project,
            nome=nome,
            cor=cor,
            descricao=descricao,
        )

        messages.success(request, f'Label "{label.nome}" criada com sucesso!')
        return redirect('roadmap_timeline') + f'?project={project_id}'

    context = {'project': project}
    return render(request, 'core/form_label.html', context)


@login_required
def editar_label(request, label_id):
    """Editar label existente"""
    from .models import Label

    label = get_object_or_404(Label, id=label_id)

    if request.method == 'POST':
        label.nome = request.POST.get('nome')
        label.cor = request.POST.get('cor', '#6b7280')
        label.descricao = request.POST.get('descricao', '')
        label.save()

        messages.success(request, f'Label "{label.nome}" atualizada com sucesso!')
        return redirect('roadmap_timeline') + f'?project={label.project.id}'

    context = {'label': label}
    return render(request, 'core/form_label.html', context)


@login_required
def excluir_label(request, label_id):
    """Excluir label"""
    from .models import Label

    label = get_object_or_404(Label, id=label_id)
    project_id = label.project.id

    if request.method == 'POST':
        nome = label.nome
        label.delete()
        messages.success(request, f'Label "{nome}" excluída com sucesso!')
        return redirect('roadmap_timeline') + f'?project={project_id}'

    return redirect('roadmap_timeline')


# --- CRUD: Tasks ---

@login_required
@require_POST
def criar_task_modal(request):
    """Criar tarefa via modal rápido (AJAX)"""
    from .models import Project, ProjectTask, TaskHistorico, Notificacao
    from django.contrib.auth.models import User
    import json

    try:
        project_id = request.POST.get('project_id')
        titulo = request.POST.get('titulo')
        data_inicio = request.POST.get('data_inicio') or None
        data_fim = request.POST.get('data_fim') or None
        milestone_id = request.POST.get('milestone_id') or None
        priority = request.POST.get('priority', 'medium')
        responsavel_ids = request.POST.getlist('responsaveis[]')  # Lista de IDs dos responsáveis

        project = get_object_or_404(Project, id=project_id)

        task = ProjectTask.objects.create(
            project=project,
            titulo=titulo,
            data_inicio=data_inicio,
            data_fim=data_fim,
            milestone_id=milestone_id,
            priority=priority,
            criado_por=request.user,
            status='todo',
        )

        # Adicionar responsáveis
        if responsavel_ids:
            for user_id in responsavel_ids:
                user = User.objects.get(id=user_id)
                task.responsaveis.add(user)

                # Criar notificação para o responsável
                Notificacao.objects.create(
                    usuario=user,
                    tipo='tarefa_atribuida',
                    titulo=f'Nova tarefa atribuída: {titulo}',
                    mensagem=f'{request.user.get_full_name() or request.user.username} atribuiu você à tarefa "{titulo}" no projeto {project.nome}.',
                    task=task,
                    lida=False
                )

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='criado',
            descricao=f'Tarefa criada: {titulo}'
        )

        return JsonResponse({
            'success': True,
            'task_id': task.id,
            'message': 'Tarefa criada com sucesso!'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao criar tarefa: {str(e)}'
        }, status=400)


@login_required
def task_detail(request, task_id):
    """Visualização detalhada da tarefa"""
    from .models import ProjectTask

    try:
        task = ProjectTask.objects.get(id=task_id)
    except ProjectTask.DoesNotExist:
        messages.warning(request, 'Esta tarefa não existe mais ou foi deletada.')
        return redirect('roadmap_timeline')

    # Buscar subtarefas
    subtasks = task.subtasks.all()

    # Histórico
    historico = task.historico.select_related('usuario').all()[:50]

    # Quantidades produzidas
    quantidades = task.quantidades_feitas.select_related('usuario').all()

    # Usuários que trabalharam
    usuarios_trabalharam = User.objects.filter(
        id__in=task.quantidades_feitas.values_list('usuario_id', flat=True).distinct()
    )

    context = {
        'task': task,
        'subtasks': subtasks,
        'historico': historico,
        'quantidades': quantidades,
        'usuarios_trabalharam': usuarios_trabalharam,
        'labels_disponiveis': task.project.labels.all(),
        'usuarios_disponiveis': User.objects.filter(is_active=True),
    }

    return render(request, 'core/task_detail.html', context)


@login_required
def editar_task(request, task_id):
    """Editar tarefa completa"""
    from .models import ProjectTask, TaskHistorico

    task = get_object_or_404(ProjectTask, id=task_id)

    if request.method == 'POST':
        # Campos básicos
        task.titulo = request.POST.get('titulo')
        task.descricao = request.POST.get('descricao', '')

        # Datas
        task.data_inicio = request.POST.get('data_inicio') or None
        task.data_fim = request.POST.get('data_fim') or None

        # Campos customizados
        task.priority = request.POST.get('priority', 'medium')
        task.status = request.POST.get('status', 'todo')
        task.estimativa = request.POST.get('estimativa') or None
        task.quantidade_meta = int(request.POST.get('quantidade_meta', 0))

        # Relações
        task.milestone_id = request.POST.get('milestone_id') or None
        task.sprint_id = request.POST.get('sprint_id') or None

        # Responsáveis (ManyToMany)
        responsaveis_ids = request.POST.getlist('responsaveis')
        task.responsaveis.set(responsaveis_ids)

        # Labels (ManyToMany)
        labels_ids = request.POST.getlist('labels')
        task.labels.set(labels_ids)

        task.save()

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='editado',
            descricao=f'Tarefa editada: {task.titulo}'
        )

        messages.success(request, f'Tarefa "{task.titulo}" atualizada com sucesso!')
        return redirect('task_detail', task_id=task.id)

    context = {
        'task': task,
        'milestones': task.project.milestones.all(),
        'sprints': task.project.sprints.all(),
        'labels': task.project.labels.all(),
        'usuarios': User.objects.filter(is_active=True),
    }

    return render(request, 'core/form_task.html', context)


@login_required
def excluir_task(request, task_id):
    """Excluir tarefa"""
    from .models import ProjectTask

    task = get_object_or_404(ProjectTask, id=task_id)
    project_id = task.project.id

    if request.method == 'POST':
        titulo = task.titulo
        task.delete()
        messages.success(request, f'Tarefa "{titulo}" excluída com sucesso!')
        from django.urls import reverse
        url = reverse('roadmap_timeline') + f'?project={project_id}'
        return redirect(url)

    return redirect('task_detail', task_id=task_id)


@login_required
@require_POST
def atualizar_task_dates(request, task_id):
    """Atualizar datas da tarefa via drag & drop no Gantt (AJAX)"""
    from .models import ProjectTask, TaskHistorico
    from datetime import datetime
    import json

    try:
        task = get_object_or_404(ProjectTask, id=task_id)

        nova_data_inicio = request.POST.get('data_inicio')
        nova_data_fim = request.POST.get('data_fim')

        data_inicio_antiga = task.data_inicio
        data_fim_antiga = task.data_fim

        if nova_data_inicio:
            task.data_inicio = datetime.strptime(nova_data_inicio, '%Y-%m-%d').date()
        if nova_data_fim:
            task.data_fim = datetime.strptime(nova_data_fim, '%Y-%m-%d').date()

        task.save()

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='prazo_alterado',
            descricao=f'Prazo alterado: {data_inicio_antiga} - {data_fim_antiga} → {task.data_inicio} - {task.data_fim}'
        )

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao atualizar datas: {str(e)}'
        }, status=400)


@login_required
@require_POST
def alterar_status(request, task_id):
    """Alterar status da tarefa (AJAX)"""
    from .models import ProjectTask, TaskHistorico
    from django.utils import timezone
    import json

    try:
        task = get_object_or_404(ProjectTask, id=task_id)

        novo_status = request.POST.get('status')
        status_anterior = task.status

        task.status = novo_status

        # Se mudou para 'done', marcar como finalizado
        if novo_status == 'done' and not task.finalizado:
            task.finalizado = True
            task.data_finalizacao = timezone.now()
        elif novo_status != 'done' and task.finalizado:
            task.finalizado = False
            task.data_finalizacao = None

        task.save()

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='editado',
            descricao=f'Status alterado: {status_anterior} → {novo_status}'
        )

        # Notificar responsáveis quando tarefa é concluída
        if novo_status == 'done' and status_anterior != 'done':
            from .models import Notificacao
            quem_concluiu = request.user.get_full_name() or request.user.username
            for responsavel in task.responsaveis.all():
                Notificacao.objects.create(
                    usuario=responsavel,
                    tipo='tarefa_concluida',
                    titulo=f'Tarefa concluída: {task.titulo}',
                    mensagem=f'{quem_concluiu} marcou a tarefa "{task.titulo}" como concluída no projeto {task.project.nome}.',
                    task=task,
                    lida=False
                )

        return JsonResponse({
            'success': True,
            'message': f'Status alterado para {task.get_status_display()}'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao alterar status: {str(e)}'
        }, status=400)


@login_required
def criar_subtask(request, task_id):
    """Criar subtarefa"""
    from .models import ProjectTask, TaskHistorico

    parent_task = get_object_or_404(ProjectTask, id=task_id)

    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descricao = request.POST.get('descricao', '')

        subtask = ProjectTask.objects.create(
            project=parent_task.project,
            parent_task=parent_task,
            titulo=titulo,
            descricao=descricao,
            criado_por=request.user,
            status='todo',
        )

        # Registrar no histórico da tarefa pai
        TaskHistorico.objects.create(
            task=parent_task,
            usuario=request.user,
            tipo_acao='subtask_criada',
            descricao=f'Subtarefa criada: {titulo}'
        )

        # Registrar no histórico da subtarefa
        TaskHistorico.objects.create(
            task=subtask,
            usuario=request.user,
            tipo_acao='criado',
            descricao=f'Subtarefa criada (pai: {parent_task.titulo})'
        )

        messages.success(request, f'Subtarefa "{titulo}" criada com sucesso!')
        return redirect('task_detail', task_id=task_id)

    context = {
        'parent_task': parent_task,
    }

    return render(request, 'core/form_subtask.html', context)


# --- Tracking de Produção (adaptado para ProjectTask) ---

@login_required
def registrar_quantidade_project(request, task_id):
    """Registrar quantidade produzida em ProjectTask"""
    from .models import ProjectTask, TaskQuantidadeFeita, TaskHistorico
    from django.utils import timezone

    task = get_object_or_404(ProjectTask, id=task_id)

    if request.method == 'POST':
        quantidade = int(request.POST.get('quantidade', 0))
        observacoes = request.POST.get('observacoes', '')

        # Criar registro de quantidade
        TaskQuantidadeFeita.objects.create(
            task=task,
            usuario=request.user,
            quantidade=quantidade,
            observacoes=observacoes,
            data=timezone.now()
        )

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='quantidade_adicionada',
            descricao=f'{request.user.get_full_name() or request.user.username} adicionou {quantidade} unidades'
        )

        messages.success(request, f'{quantidade} unidades registradas com sucesso!')
        return redirect('task_detail', task_id=task.id)

    context = {'task': task}
    return render(request, 'core/form_quantidade.html', context)


@login_required
def editar_quantidade_project(request, quantidade_id):
    """Editar quantidade produzida"""
    from .models import TaskQuantidadeFeita, TaskHistorico

    quantidade_obj = get_object_or_404(TaskQuantidadeFeita, id=quantidade_id)
    task = quantidade_obj.task

    if request.method == 'POST':
        quantidade_antiga = quantidade_obj.quantidade
        quantidade_obj.quantidade = int(request.POST.get('quantidade', 0))
        quantidade_obj.observacoes = request.POST.get('observacoes', '')
        quantidade_obj.save()

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='editado',
            descricao=f'Quantidade alterada: {quantidade_antiga} → {quantidade_obj.quantidade}'
        )

        messages.success(request, 'Quantidade atualizada com sucesso!')
        return redirect('task_detail', task_id=task.id)

    context = {
        'quantidade': quantidade_obj,
        'task': task,
    }

    return render(request, 'core/form_quantidade.html', context)


@login_required
def excluir_quantidade_project(request, quantidade_id):
    """Excluir quantidade produzida"""
    from .models import TaskQuantidadeFeita, TaskHistorico

    quantidade_obj = get_object_or_404(TaskQuantidadeFeita, id=quantidade_id)
    task = quantidade_obj.task

    if request.method == 'POST':
        quantidade_valor = quantidade_obj.quantidade
        quantidade_obj.delete()

        # Registrar no histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='editado',
            descricao=f'Registro de quantidade removido: {quantidade_valor} unidades'
        )

        messages.success(request, 'Quantidade removida com sucesso!')
        return redirect('task_detail', task_id=task.id)

    return redirect('task_detail', task_id=task.id)


# --- Views do Kanban ---

@login_required
def kanban_board(request):
    colunas = KanbanColumn.objects.all().order_by('ordem')
    tarefas_por_coluna = {str(coluna.id): list(coluna.tasks.all().order_by('ordem')) for coluna in colunas}
    historico = TaskQuantidadeFeita.objects.select_related('usuario', 'task').order_by('-data')[:30]
    contexto = {
        'colunas': colunas,
        'tarefas_por_coluna': tarefas_por_coluna,
        'historico': historico,
    }
    return render(request, 'core/kanban_board.html', contexto)

@login_required
def detalhe_tarefa(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    historico = task.historico.all()
    quantidades = task.quantidades_feitas.all()

    # Usuários que trabalharam neste card
    usuarios_trabalharam = User.objects.filter(
        Q(taskhistorico__task=task) | Q(taskquantidadefeita__task=task)
    ).distinct()

    contexto = {
        'task': task,
        'historico': historico,
        'quantidades': quantidades,
        'usuarios_trabalharam': usuarios_trabalharam,
    }
    return render(request, 'core/detalhe_tarefa.html', contexto)

@login_required
def metricas_kanban(request):
    from django.db.models import Count, Sum, Q

    # Cards finalizados por usuário - conta apenas cards ATUALMENTE finalizados
    # Para cada usuário, busca o último histórico de finalização/reabertura de cada card
    # e conta apenas se o último registro for de finalização

    # Cálculo proporcional: cada usuário recebe crédito fracionado baseado em sua contribuição
    # Exemplo: Card de 100 unidades - User A produziu 60 (0.6 cards) + User B produziu 40 (0.4 cards)

    from decimal import Decimal

    # Primeiro, pegamos todos os usuários que produziram algo
    usuarios = User.objects.annotate(
        total_quantidade_produzida=Sum('taskquantidadefeita__quantidade')
    ).filter(total_quantidade_produzida__gt=0)

    # Para cada usuário, calculamos sua contribuição proporcional em cards finalizados
    cards_por_usuario = []
    for user in usuarios:
        # Buscar todos os cards finalizados atualmente
        tasks_finalizadas = Task.objects.filter(finalizado=True)

        cards_finalizados_proporcional = Decimal('0.0')

        # Para cada card finalizado, calcular a contribuição proporcional do usuário
        for task in tasks_finalizadas:
            # Total produzido neste card
            total_produzido_card = task.quantidade_produzida

            if total_produzido_card > 0:
                # Quantidade que este usuário produziu neste card
                quantidade_usuario = TaskQuantidadeFeita.objects.filter(
                    task=task,
                    usuario=user
                ).aggregate(total=Sum('quantidade'))['total'] or 0

                # Calcular proporção (0.0 a 1.0)
                if quantidade_usuario > 0:
                    proporcao = Decimal(quantidade_usuario) / Decimal(total_produzido_card)
                    cards_finalizados_proporcional += proporcao

        if cards_finalizados_proporcional > 0 or user.total_quantidade_produzida:
            media_por_card = (user.total_quantidade_produzida / float(cards_finalizados_proporcional)) if cards_finalizados_proporcional > 0 else 0
            cards_por_usuario.append({
                'user': user,
                'total_finalizados': float(cards_finalizados_proporcional),
                'total_quantidade_produzida': user.total_quantidade_produzida or 0,
                'media_por_card': media_por_card
            })

    # Ordenar por total_finalizados (decrescente)
    cards_por_usuario = sorted(cards_por_usuario, key=lambda x: x['total_finalizados'], reverse=True)

    # Total de cards finalizados ATUALMENTE
    total_cards_finalizados = Task.objects.filter(finalizado=True).count()

    # Total de cards em andamento
    total_cards_andamento = Task.objects.filter(em_andamento=True, finalizado=False).count()

    # Total de quantidade produzida
    total_produzido = TaskQuantidadeFeita.objects.aggregate(total=Sum('quantidade'))['total'] or 0

    contexto = {
        'cards_por_usuario': cards_por_usuario,
        'total_cards_finalizados': total_cards_finalizados,
        'total_cards_andamento': total_cards_andamento,
        'total_produzido': total_produzido,
    }
    return render(request, 'core/metricas_kanban.html', contexto)

@login_required
def criar_coluna(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        cor = request.POST.get('cor', '#f3f4f6')
        ordem = KanbanColumn.objects.count()
        coluna = KanbanColumn.objects.create(nome=nome, cor=cor, ordem=ordem)
        return redirect('kanban_board')
    return render(request, 'core/form_coluna.html')

@login_required
def editar_coluna(request, coluna_id):
    coluna = get_object_or_404(KanbanColumn, id=coluna_id)
    if request.method == 'POST':
        coluna.nome = request.POST.get('nome')
        coluna.cor = request.POST.get('cor', coluna.cor)

        # Processar mudança de posição
        nova_posicao = request.POST.get('posicao')
        if nova_posicao is not None:
            nova_posicao = int(nova_posicao)
            posicao_atual = coluna.ordem

            if nova_posicao != posicao_atual:
                # Reordenar todas as colunas
                colunas = list(KanbanColumn.objects.all().order_by('ordem'))

                # Remove a coluna atual da lista
                colunas.pop(posicao_atual)

                # Insere na nova posição
                colunas.insert(nova_posicao, coluna)

                # Atualiza a ordem de todas as colunas
                for idx, col in enumerate(colunas):
                    col.ordem = idx
                    col.save(update_fields=['ordem'])

        coluna.save()
        return redirect('kanban_board')

    # Preparar dados para o template
    total_colunas = KanbanColumn.objects.count()
    context = {
        'coluna': coluna,
        'range_posicoes': range(total_colunas),
        'posicao_atual': coluna.ordem,
        'total_colunas': total_colunas
    }
    return render(request, 'core/form_coluna.html', context)

@login_required
def excluir_coluna(request, coluna_id):
    coluna = get_object_or_404(KanbanColumn, id=coluna_id)
    if request.method == 'POST':
        coluna.delete()
        return redirect('kanban_board')
    return render(request, 'core/confirmar_exclusao.html', {'obj': coluna, 'tipo': 'coluna'})

@login_required
def criar_tarefa(request, coluna_id=None):
    if request.method == 'POST':
        coluna = get_object_or_404(KanbanColumn, id=request.POST.get('coluna_id', coluna_id))
        titulo = request.POST.get('titulo')
        descricao = request.POST.get('descricao', '')
        quantidade_meta = int(request.POST.get('quantidade', 0))
        ordem = coluna.tasks.count()
        task = Task.objects.create(coluna=coluna, titulo=titulo, descricao=descricao, quantidade_meta=quantidade_meta, ordem=ordem)
        responsaveis_ids = request.POST.getlist('responsaveis')
        if responsaveis_ids:
            task.responsaveis.set(responsaveis_ids)

        # Registra histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='criado',
            descricao=f'Card criado na coluna "{coluna.nome}" com meta de {quantidade_meta} unidades'
        )
        return redirect('kanban_board')
    colunas = KanbanColumn.objects.all()
    users = User.objects.all()
    return render(request, 'core/form_tarefa.html', {'colunas': colunas, 'users': users, 'coluna_id': coluna_id})

@login_required
def editar_tarefa(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    if request.method == 'POST':
        old_titulo = task.titulo
        old_responsaveis = set(task.responsaveis.all())

        task.titulo = request.POST.get('titulo')
        task.descricao = request.POST.get('descricao', '')
        task.quantidade_meta = int(request.POST.get('quantidade', 0))
        task.coluna_id = request.POST.get('coluna_id', task.coluna_id)
        responsaveis_ids = request.POST.getlist('responsaveis')
        task.responsaveis.set(responsaveis_ids)
        task.save()

        new_responsaveis = set(task.responsaveis.all())

        # Registra histórico de edição
        if old_titulo != task.titulo:
            TaskHistorico.objects.create(
                task=task,
                usuario=request.user,
                tipo_acao='editado',
                descricao=f'Título alterado de "{old_titulo}" para "{task.titulo}"'
            )

        # Verifica mudanças nos responsáveis
        adicionados = new_responsaveis - old_responsaveis
        removidos = old_responsaveis - new_responsaveis

        for user in adicionados:
            TaskHistorico.objects.create(
                task=task,
                usuario=request.user,
                tipo_acao='responsavel_adicionado',
                descricao=f'{user.get_full_name() or user.username} adicionado como responsável'
            )

        for user in removidos:
            TaskHistorico.objects.create(
                task=task,
                usuario=request.user,
                tipo_acao='responsavel_removido',
                descricao=f'{user.get_full_name() or user.username} removido dos responsáveis'
            )

        return redirect('kanban_board')
    colunas = KanbanColumn.objects.all()
    users = User.objects.all()
    return render(request, 'core/form_tarefa.html', {'task': task, 'colunas': colunas, 'users': users})

@login_required
def excluir_tarefa(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    if request.method == 'POST':
        task.delete()
        return redirect('kanban_board')
    return render(request, 'core/confirmar_exclusao.html', {'obj': task, 'tipo': 'tarefa'})

@login_required
@require_POST
def mover_tarefa(request, task_id):
    try:
        task = get_object_or_404(Task, id=task_id)
        nova_coluna_id = request.POST.get('nova_coluna_id')
        nova_ordem = int(request.POST.get('nova_ordem', 0))

        if nova_coluna_id:
            coluna_antiga = task.coluna
            nova_coluna = get_object_or_404(KanbanColumn, id=nova_coluna_id)

            if coluna_antiga.id != nova_coluna.id:
                task.coluna = nova_coluna
                # Registra histórico
                TaskHistorico.objects.create(
                    task=task,
                    usuario=request.user,
                    tipo_acao='movido',
                    descricao=f'Movido de "{coluna_antiga.nome}" para "{nova_coluna.nome}"'
                )

        task.ordem = nova_ordem
        task.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def mover_coluna(request, coluna_id):
    coluna = get_object_or_404(KanbanColumn, id=coluna_id)
    nova_ordem = int(request.POST.get('nova_ordem', 0))
    coluna.ordem = nova_ordem
    coluna.save()
    return JsonResponse({'success': True})

@login_required
@require_POST
def marcar_andamento(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    task.em_andamento = True
    task.save()

    # Registra histórico
    TaskHistorico.objects.create(
        task=task,
        usuario=request.user,
        tipo_acao='iniciado',
        descricao=f'Card iniciado por {request.user.get_full_name() or request.user.username}'
    )
    return redirect('kanban_board')

@login_required
@require_POST
def finalizar(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    task.em_andamento = False
    task.finalizado = True
    task.data_finalizacao = timezone.now()
    task.save()

    # Registra histórico
    TaskHistorico.objects.create(
        task=task,
        usuario=request.user,
        tipo_acao='finalizado',
        descricao=f'Card finalizado por {request.user.get_full_name() or request.user.username}'
    )
    return redirect('kanban_board')

@login_required
@require_POST
def desfinalizar(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    task.finalizado = False
    task.data_finalizacao = None
    task.save()

    # Registra histórico
    TaskHistorico.objects.create(
        task=task,
        usuario=request.user,
        tipo_acao='editado',
        descricao=f'Card desfinalizadopor {request.user.get_full_name() or request.user.username}'
    )
    return redirect('kanban_board')

@login_required
@require_POST
def registrar_quantidade(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    quantidade = int(request.POST.get('quantidade', 0))
    if quantidade > 0:
        TaskQuantidadeFeita.objects.create(task=task, usuario=request.user, quantidade=quantidade)

        # Registra histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='quantidade_adicionada',
            descricao=f'{request.user.get_full_name() or request.user.username} produziu {quantidade} unidade(s). Total: {task.quantidade_produzida}/{task.quantidade_meta}'
        )
    return redirect('kanban_board')

@login_required
@require_POST
def editar_quantidade_feita(request, quantidade_id):
    """Edita uma quantidade feita existente"""
    quantidade_obj = get_object_or_404(TaskQuantidadeFeita, id=quantidade_id)
    task = quantidade_obj.task
    quantidade_antiga = quantidade_obj.quantidade

    nova_quantidade = int(request.POST.get('quantidade', 0))
    if nova_quantidade > 0:
        quantidade_obj.quantidade = nova_quantidade
        quantidade_obj.save()

        # Registra histórico
        TaskHistorico.objects.create(
            task=task,
            usuario=request.user,
            tipo_acao='quantidade_editada',
            descricao=f'{request.user.get_full_name() or request.user.username} editou quantidade de {quantidade_antiga} para {nova_quantidade} unidades. Total: {task.quantidade_produzida}/{task.quantidade_meta}'
        )
        messages.success(request, f'Quantidade atualizada de {quantidade_antiga} para {nova_quantidade} unidades!')
    else:
        messages.error(request, 'Quantidade deve ser maior que zero.')

    return redirect('kanban_board')

@login_required
@require_POST
def excluir_quantidade_feita(request, quantidade_id):
    """Exclui uma quantidade feita"""
    quantidade_obj = get_object_or_404(TaskQuantidadeFeita, id=quantidade_id)
    task = quantidade_obj.task
    quantidade_valor = quantidade_obj.quantidade
    usuario_nome = quantidade_obj.usuario.get_full_name() or quantidade_obj.usuario.username if quantidade_obj.usuario else 'Usuário desconhecido'

    # Registra histórico antes de excluir
    TaskHistorico.objects.create(
        task=task,
        usuario=request.user,
        tipo_acao='quantidade_removida',
        descricao=f'{request.user.get_full_name() or request.user.username} removeu registro de {quantidade_valor} unidades (feito por {usuario_nome}). Total atual: {task.quantidade_produzida - quantidade_valor}/{task.quantidade_meta}'
    )

    quantidade_obj.delete()
    messages.success(request, f'Registro de {quantidade_valor} unidades removido com sucesso!')

    return redirect('kanban_board')

# --- Views de Controle de Ponto ---

@login_required
def controle_ponto(request):
    """View principal do controle de ponto"""
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    import calendar

    # Verifica se é superusuário ou o próprio usuário
    usuario_id = request.GET.get('usuario_id')
    if usuario_id and request.user.is_superuser:
        usuario = get_object_or_404(User, id=usuario_id)
    else:
        usuario = request.user

    # Criar jornada se não existir
    jornada, created = JornadaTrabalho.objects.get_or_create(
        usuario=usuario,
        defaults={'horas_diarias': 9.0, 'horas_sexta': 8.0, 'intervalo_almoco': 1.0}
    )

    # Dados do mês atual (usando período personalizado do funcionário)
    now = datetime.now()
    dia_inicio = jornada.dia_inicio_mes
    dia_fim = jornada.dia_fim_mes

    # Calcular primeiro e último dia do período personalizado
    if dia_inicio == 1 and dia_fim == 0:
        # Período padrão: do dia 1 ao último dia do mês
        primeiro_dia = datetime(now.year, now.month, 1)
        ultimo_dia = datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1], 23, 59, 59)
    else:
        # Período personalizado
        if dia_fim == 0:
            # 0 significa último dia do mês
            dia_fim = calendar.monthrange(now.year, now.month)[1]

        # Se estamos depois do dia de início, o período é do dia_inicio deste mês até dia_fim do próximo mês
        # Se estamos antes do dia de início, o período é do dia_inicio do mês passado até dia_fim deste mês
        if now.day >= dia_inicio:
            # Período atual: dia_inicio deste mês até dia_fim do próximo mês (ou mesmo mês se dia_fim > dia_inicio)
            primeiro_dia = datetime(now.year, now.month, dia_inicio, 0, 0, 0)

            if dia_fim >= dia_inicio:
                # Período termina no mesmo mês
                ultimo_dia = datetime(now.year, now.month, dia_fim, 23, 59, 59)
            else:
                # Período termina no próximo mês
                if now.month == 12:
                    proximo_mes = 1
                    proximo_ano = now.year + 1
                else:
                    proximo_mes = now.month + 1
                    proximo_ano = now.year
                ultimo_dia = datetime(proximo_ano, proximo_mes, dia_fim, 23, 59, 59)
        else:
            # Período anterior: dia_inicio do mês passado até dia_fim deste mês
            if now.month == 1:
                mes_anterior = 12
                ano_anterior = now.year - 1
            else:
                mes_anterior = now.month - 1
                ano_anterior = now.year

            primeiro_dia = datetime(ano_anterior, mes_anterior, dia_inicio, 0, 0, 0)
            ultimo_dia = datetime(now.year, now.month, dia_fim, 23, 59, 59)

    # Registros de ponto do mês
    pontos_mes = RegistroPonto.objects.filter(
        usuario=usuario,
        data_hora__gte=primeiro_dia,
        data_hora__lte=ultimo_dia
    ).order_by('-data_hora')

    # Calcular horas trabalhadas no mês
    horas_trabalhadas = 0
    dias_trabalhados = {}

    for ponto in pontos_mes:
        dia = ponto.data_hora.date()
        if dia not in dias_trabalhados:
            dias_trabalhados[dia] = {'entrada': None, 'saida': None, 'inicio_almoco': None, 'fim_almoco': None}

        if ponto.tipo == 'entrada':
            dias_trabalhados[dia]['entrada'] = ponto.data_hora
        elif ponto.tipo == 'saida':
            dias_trabalhados[dia]['saida'] = ponto.data_hora
        elif ponto.tipo == 'inicio_almoco':
            dias_trabalhados[dia]['inicio_almoco'] = ponto.data_hora
        elif ponto.tipo == 'fim_almoco':
            dias_trabalhados[dia]['fim_almoco'] = ponto.data_hora

    # Calcular horas de cada dia
    for dia, registros in dias_trabalhados.items():
        if registros['entrada'] and registros['saida']:
            delta = registros['saida'] - registros['entrada']
            horas_brutas = delta.total_seconds() / 3600

            # Descontar intervalo de almoço se registrado
            if registros['inicio_almoco'] and registros['fim_almoco']:
                intervalo = registros['fim_almoco'] - registros['inicio_almoco']
                horas_brutas -= intervalo.total_seconds() / 3600

            horas_trabalhadas += horas_brutas

    # Calcular horas abonadas separadamente
    abonos_periodo = AbonoDia.objects.filter(
        usuario=usuario,
        data__gte=primeiro_dia.date(),
        data__lte=ultimo_dia.date()
    )
    horas_abonadas = sum(float(abono.horas_abonadas) for abono in abonos_periodo)

    # Pesquisa por dia específico
    dia_pesquisa_str = request.GET.get('dia_pesquisa')
    dia_pesquisado = None
    pontos_do_dia = []
    entrada = None
    saida = None
    inicio_almoco = None
    fim_almoco = None
    horas_trabalhadas_dia = 0
    tempo_almoco = None

    if dia_pesquisa_str:
        try:
            dia_pesquisado = datetime.strptime(dia_pesquisa_str, '%Y-%m-%d').date()
            # Buscar todos os pontos do dia
            pontos_do_dia = RegistroPonto.objects.filter(
                usuario=usuario,
                data_hora__date=dia_pesquisado
            ).order_by('data_hora')

            # Separar os registros por tipo
            registros_dia = {'entrada': None, 'saida': None, 'inicio_almoco': None, 'fim_almoco': None}
            for ponto in pontos_do_dia:
                if ponto.tipo == 'entrada':
                    entrada = registros_dia['entrada'] = ponto.data_hora
                elif ponto.tipo == 'saida':
                    saida = registros_dia['saida'] = ponto.data_hora
                elif ponto.tipo == 'inicio_almoco':
                    inicio_almoco = registros_dia['inicio_almoco'] = ponto.data_hora
                elif ponto.tipo == 'fim_almoco':
                    fim_almoco = registros_dia['fim_almoco'] = ponto.data_hora

            # Calcular horas trabalhadas do dia
            if registros_dia['entrada'] and registros_dia['saida']:
                delta = registros_dia['saida'] - registros_dia['entrada']
                horas_trabalhadas_dia = delta.total_seconds() / 3600

                # Calcular tempo de almoço
                if registros_dia['inicio_almoco'] and registros_dia['fim_almoco']:
                    intervalo = registros_dia['fim_almoco'] - registros_dia['inicio_almoco']
                    tempo_almoco = intervalo.total_seconds() / 3600
                    horas_trabalhadas_dia -= tempo_almoco
        except (ValueError, TypeError):
            # Data inválida, ignora
            dia_pesquisado = None

    # Calcular dias de falta (dias úteis sem ponto e sem abono, até hoje)
    dias_falta = 0
    dia_atual = primeiro_dia.date()
    hoje = now.date()
    dias_semana_trabalho = [int(d) for d in jornada.dias_semana.split(',')]

    # Calcular apenas até hoje, não até o fim do período
    data_limite = min(ultimo_dia.date(), hoje)

    while dia_atual <= data_limite:
        # Verifica se é dia de trabalho (considerando dias_semana configurados)
        if dia_atual.weekday() in dias_semana_trabalho:
            # Verifica se não tem registro de ponto neste dia
            tem_ponto = RegistroPonto.objects.filter(
                usuario=usuario,
                data_hora__date=dia_atual
            ).exists()

            # Verifica se não tem abono neste dia
            tem_abono = AbonoDia.objects.filter(
                usuario=usuario,
                data=dia_atual
            ).exists()

            # Se não tem ponto e não tem abono, é falta
            if not tem_ponto and not tem_abono:
                dias_falta += 1

        dia_atual += timedelta(days=1)

    # Último ponto do dia
    ultimo_ponto_hoje = RegistroPonto.objects.filter(
        usuario=usuario,
        data_hora__date=hoje
    ).order_by('-data_hora').first()

    # Meta mensal (horas esperadas no período completo)
    horas_meta_mensal = jornada.horas_esperadas_periodo(primeiro_dia.date(), ultimo_dia.date())

    # Horas esperadas até hoje (para cálculo do saldo)
    data_final_calculo = min(ultimo_dia.date(), hoje)
    horas_esperadas = jornada.horas_esperadas_periodo(primeiro_dia.date(), data_final_calculo)

    # Saldo de horas (inclui horas abonadas no cálculo)
    saldo_horas = (horas_trabalhadas + horas_abonadas) - horas_esperadas

    # Últimos 30 dias para gráfico com cores baseadas em horas trabalhadas
    trinta_dias_atras = now - timedelta(days=30)
    presenca_ultimos_30 = []

    # Obter dias de trabalho configurados na jornada
    dias_semana_trabalho = [int(d) for d in jornada.dias_semana.split(',')]

    for i in range(30):
        dia = (now - timedelta(days=29-i)).date()

        # Buscar pontos do dia
        pontos_dia = RegistroPonto.objects.filter(
            usuario=usuario,
            data_hora__date=dia
        ).order_by('data_hora')

        # Calcular horas trabalhadas no dia
        horas_dia = 0
        registros_dia = {'entrada': None, 'saida': None, 'inicio_almoco': None, 'fim_almoco': None}

        for ponto in pontos_dia:
            if ponto.tipo == 'entrada':
                registros_dia['entrada'] = ponto.data_hora
            elif ponto.tipo == 'saida':
                registros_dia['saida'] = ponto.data_hora
            elif ponto.tipo == 'inicio_almoco':
                registros_dia['inicio_almoco'] = ponto.data_hora
            elif ponto.tipo == 'fim_almoco':
                registros_dia['fim_almoco'] = ponto.data_hora

        # Verificar se o dia foi abonado
        abono_dia = AbonoDia.objects.filter(usuario=usuario, data=dia).first()

        # Calcular horas líquidas
        if registros_dia['entrada'] and registros_dia['saida']:
            delta = registros_dia['saida'] - registros_dia['entrada']
            horas_dia = delta.total_seconds() / 3600

            # Descontar intervalo de almoço
            if registros_dia['inicio_almoco'] and registros_dia['fim_almoco']:
                intervalo = registros_dia['fim_almoco'] - registros_dia['inicio_almoco']
                horas_dia -= intervalo.total_seconds() / 3600

        # Definir cor baseada no percentual das horas esperadas (considera sexta diferente)
        horas_esperadas_dia = jornada.horas_esperadas_dia(dia)

        # Verificar se é dia útil (está nos dias de trabalho configurados)
        dia_util = dia.weekday() in dias_semana_trabalho

        # Se o dia foi abonado, usar cor amarela
        if abono_dia:
            cor = 'bg-yellow-500/50'  # Amarelo com 50% transparência: dia abonado
            percentual = 100  # Considerar como 100% por ser abonado
            horas_dia = float(abono_dia.horas_abonadas)  # Usar horas abonadas
        elif not dia_util:
            cor = 'bg-gray-300'  # Cinza: dia não útil (fim de semana/feriado)
            percentual = 0
        elif horas_dia == 0:
            cor = 'bg-red-500/50'  # Vermelho com 50% transparência: falta (dia útil sem registro)
            percentual = 0
        else:
            percentual = (horas_dia / horas_esperadas_dia) * 100
            if percentual >= 100:
                cor = 'bg-green-500'  # Verde chamativo: 100%+
            elif percentual >= 75:
                cor = 'bg-green-500/75'  # Verde com 25% transparência: 75-99%
            elif percentual >= 50:
                cor = 'bg-green-500/55'  # Verde com 45% transparência: 50-74%
            elif percentual >= 25:
                cor = 'bg-green-500/35'  # Verde com 65% transparência: 25-49%
            else:
                cor = 'bg-green-500/15'  # Verde com 85% transparência: 1-24%

        presenca_ultimos_30.append({
            'dia': dia.strftime('%d/%m'),
            'presente': horas_dia > 0 or abono_dia is not None,
            'horas': round(horas_dia, 2),
            'cor': cor,
            'percentual': round(percentual, 0),
            'abonado': abono_dia is not None,
            'tipo_abono': abono_dia.get_tipo_abono_display() if abono_dia else None,
            'dia_util': dia_util
        })

    # Lista de usuários (apenas para superusuário)
    usuarios = User.objects.filter(is_active=True) if request.user.is_superuser else []

    # Abonos do mês (para superusuário)
    abonos_mes = AbonoDia.objects.filter(
        data__gte=primeiro_dia.date(),
        data__lte=ultimo_dia.date()
    ).order_by('-data') if request.user.is_superuser else []

    contexto = {
        'usuario_selecionado': usuario,
        'jornada': jornada,
        'pontos_mes': pontos_mes[:10],  # Últimos 10 registros
        'horas_trabalhadas': round(horas_trabalhadas, 2),
        'horas_abonadas': round(horas_abonadas, 2),
        'horas_esperadas': round(horas_esperadas, 2),
        'horas_meta_mensal': round(horas_meta_mensal, 2),
        'saldo_horas': round(saldo_horas, 2),
        'dias_trabalhados': len([d for d, r in dias_trabalhados.items() if r['entrada'] and r['saida']]),
        'dias_falta': dias_falta,
        'ultimo_ponto_hoje': ultimo_ponto_hoje,
        'presenca_ultimos_30': presenca_ultimos_30,
        'usuarios': usuarios,
        'abonos_mes': abonos_mes,
        'primeiro_dia': primeiro_dia,
        'ultimo_dia': ultimo_dia,
        # Dados da pesquisa por dia
        'dia_pesquisado': dia_pesquisado,
        'pontos_do_dia': pontos_do_dia,
        'entrada': entrada,
        'saida': saida,
        'inicio_almoco': inicio_almoco,
        'fim_almoco': fim_almoco,
        'horas_trabalhadas_dia': round(horas_trabalhadas_dia, 2) if horas_trabalhadas_dia else 0,
        'tempo_almoco': round(tempo_almoco, 2) if tempo_almoco else None,
    }

    return render(request, 'core/controle_ponto.html', contexto)

@login_required
@require_POST
def bater_ponto(request):
    """Registra entrada, saída ou almoço"""
    tipo = request.POST.get('tipo')
    observacao = request.POST.get('observacao', '')

    if tipo not in ['entrada', 'saida', 'inicio_almoco', 'fim_almoco']:
        messages.error(request, 'Tipo de ponto inválido.')
        return redirect('controle_ponto')

    # Verifica o último ponto do dia
    hoje = timezone.now().date()
    ultimo_ponto = RegistroPonto.objects.filter(
        usuario=request.user,
        data_hora__date=hoje
    ).order_by('-data_hora').first()

    # Validações
    if tipo == 'entrada' and ultimo_ponto and ultimo_ponto.tipo in ['entrada', 'inicio_almoco']:
        messages.error(request, 'Você já registrou uma entrada. Registre a saída ou fim do almoço primeiro.')
        return redirect('controle_ponto')

    if tipo == 'inicio_almoco' and (not ultimo_ponto or ultimo_ponto.tipo != 'entrada'):
        messages.error(request, 'Você precisa registrar a entrada antes de iniciar o almoço.')
        return redirect('controle_ponto')

    if tipo == 'fim_almoco' and (not ultimo_ponto or ultimo_ponto.tipo != 'inicio_almoco'):
        messages.error(request, 'Você precisa registrar o início do almoço primeiro.')
        return redirect('controle_ponto')

    if tipo == 'saida' and (not ultimo_ponto or ultimo_ponto.tipo in ['saida', 'inicio_almoco']):
        messages.error(request, 'Você precisa registrar entrada/fim de almoço antes da saída.')
        return redirect('controle_ponto')

    # Registra o ponto
    RegistroPonto.objects.create(
        usuario=request.user,
        tipo=tipo,
        observacao=observacao
    )

    # Mensagem de sucesso
    mensagens_tipo = {
        'entrada': 'Entrada',
        'saida': 'Saída',
        'inicio_almoco': 'Início do Almoço',
        'fim_almoco': 'Fim do Almoço'
    }
    tipo_texto = mensagens_tipo.get(tipo, tipo)
    messages.success(request, f'{tipo_texto} registrado com sucesso às {timezone.now().strftime("%H:%M")}!')

    return redirect('controle_ponto')

@login_required
def abonar_dia(request):
    """Permite superuser abonar dia completo de um funcionário"""
    if not request.user.is_superuser:
        messages.error(request, 'Você não tem permissão para abonar dias.')
        return redirect('controle_ponto')

    if request.method == 'POST':
        from datetime import datetime
        usuario_id = request.POST.get('usuario_id')
        data_str = request.POST.get('data')
        tipo_abono = request.POST.get('tipo_abono')
        motivo = request.POST.get('motivo')
        horas_abonadas = request.POST.get('horas_abonadas', 8.0)

        try:
            usuario = User.objects.get(pk=usuario_id)
            data = datetime.strptime(data_str, '%Y-%m-%d').date()

            # Verificar se já existe abono para este dia
            if AbonoDia.objects.filter(usuario=usuario, data=data).exists():
                messages.warning(request, f'Já existe um abono para {usuario.username} no dia {data.strftime("%d/%m/%Y")}.')
                return redirect('controle_ponto')

            # Criar abono
            AbonoDia.objects.create(
                usuario=usuario,
                data=data,
                tipo_abono=tipo_abono,
                motivo=motivo,
                horas_abonadas=horas_abonadas,
                abonado_por=request.user
            )

            messages.success(request, f'Dia {data.strftime("%d/%m/%Y")} abonado com sucesso para {usuario.username}!')
            return redirect('controle_ponto')

        except User.DoesNotExist:
            messages.error(request, 'Usuário não encontrado.')
        except ValueError:
            messages.error(request, 'Data inválida.')
        except Exception as e:
            messages.error(request, f'Erro ao abonar dia: {str(e)}')

    return redirect('controle_ponto')

@superuser_required
def remover_abono_dia(request, abono_id):
    """Permite superuser remover um abono de dia"""

    try:
        abono = AbonoDia.objects.get(pk=abono_id)
        usuario_nome = abono.usuario.username
        data = abono.data.strftime("%d/%m/%Y")
        abono.delete()
        messages.success(request, f'Abono do dia {data} de {usuario_nome} foi removido.')
    except AbonoDia.DoesNotExist:
        messages.error(request, 'Abono não encontrado.')

    return redirect('controle_ponto')

@superuser_required
def configurar_periodo_mes(request):
    """Permite superuser configurar período de mês personalizado para funcionário"""

    if request.method == 'POST':
        usuario_id = request.POST.get('usuario_id')
        dia_inicio = request.POST.get('dia_inicio_mes')
        dia_fim = request.POST.get('dia_fim_mes')

        try:
            usuario = User.objects.get(pk=usuario_id)
            jornada, created = JornadaTrabalho.objects.get_or_create(usuario=usuario)

            # Validar dias (1-31 ou 0 para último dia)
            dia_inicio = int(dia_inicio)
            dia_fim = int(dia_fim)

            if dia_inicio < 1 or dia_inicio > 31:
                messages.error(request, 'Dia de início deve ser entre 1 e 31.')
                return redirect('controle_ponto')

            if dia_fim < 0 or dia_fim > 31:
                messages.error(request, 'Dia de fim deve ser entre 0 (último dia) e 31.')
                return redirect('controle_ponto')

            # Atualizar jornada
            jornada.dia_inicio_mes = dia_inicio
            jornada.dia_fim_mes = dia_fim
            jornada.save()

            if dia_fim == 0:
                periodo_info = f"dia {dia_inicio} ao último dia do mês"
            else:
                periodo_info = f"dia {dia_inicio} ao dia {dia_fim}"

            messages.success(request, f'Período personalizado configurado para {usuario.username}: {periodo_info}')
            return redirect('controle_ponto')

        except User.DoesNotExist:
            messages.error(request, 'Usuário não encontrado.')
        except ValueError:
            messages.error(request, 'Valores inválidos para os dias.')
        except Exception as e:
            messages.error(request, f'Erro ao configurar período: {str(e)}')

    return redirect('controle_ponto')

# === VIEWS DE CLIENTES ===

@login_required
def lista_clientes(request):
    from django.db.models import Count

    # SEGURANÇA: Filtrar apenas clientes da empresa do usuário
    clientes = Cliente.objects.all()
    clientes = filter_by_empresa(clientes, request.user)
    clientes = clientes.annotate(total_produtos=Count('produtos_fornecidos')).order_by('nome')

    # Filtros
    query = request.GET.get('q', '')
    mercado = request.GET.get('mercado', '')

    if query:
        clientes = clientes.filter(Q(nome__icontains=query) | Q(email__icontains=query))
    if mercado:
        clientes = clientes.filter(mercado=mercado)

    contexto = {
        'clientes': clientes,
        'query': query,
        'mercado_filtro': mercado,
        'total_clientes': clientes.count(),
    }
    return render(request, 'core/lista_clientes.html', contexto)

@login_required
def adicionar_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save(commit=False)
            primeira_empresa = Empresa.objects.first()
            if primeira_empresa:
                cliente.empresa = primeira_empresa
            cliente.save()
            form.save_m2m()
            messages.success(request, f'Cliente "{cliente.nome}" cadastrado com sucesso!')
            return redirect('lista_clientes')
    else:
        form = ClienteForm()

    return render(request, 'core/form_cliente.html', {'form': form, 'titulo': 'Adicionar Cliente'})

@login_required
def editar_cliente(request, pk):
    # SEGURANÇA: Garantir que o cliente pertence à empresa do usuário
    empresas = get_user_empresa(request.user)
    cliente = get_object_or_404(Cliente, pk=pk, empresa__in=empresas)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente "{cliente.nome}" atualizado com sucesso!')
            return redirect('detalhe_cliente', pk=pk)
    else:
        form = ClienteForm(instance=cliente)

    return render(request, 'core/form_cliente.html', {'form': form, 'titulo': f'Editar {cliente.nome}', 'cliente': cliente})

@login_required
def detalhe_cliente(request, pk):
    # SEGURANÇA: Garantir que o cliente pertence à empresa do usuário
    empresas = get_user_empresa(request.user)
    cliente = get_object_or_404(Cliente, pk=pk, empresa__in=empresas)
    return render(request, 'core/detalhe_cliente.html', {'cliente': cliente})

@login_required
def excluir_cliente(request, pk):
    # SEGURANÇA: Garantir que o cliente pertence à empresa do usuário
    empresas = get_user_empresa(request.user)
    cliente = get_object_or_404(Cliente, pk=pk, empresa__in=empresas)

    if request.method == 'POST':
        nome = cliente.nome
        cliente.delete()
        messages.success(request, f'Cliente "{nome}" excluído com sucesso!')
        return redirect('lista_clientes')
    return redirect('detalhe_cliente', pk=pk)

# === VIEWS DE FORNECEDORES ===

@login_required
def lista_fornecedores(request):
    # SEGURANÇA: Filtrar apenas fornecedores da empresa do usuário
    fornecedores = Fornecedor.objects.all()
    fornecedores = filter_by_empresa(fornecedores, request.user)
    fornecedores = fornecedores.order_by('nome')

    # Filtros
    query = request.GET.get('q', '')
    mercado = request.GET.get('mercado', '')

    if query:
        fornecedores = fornecedores.filter(Q(nome__icontains=query) | Q(email__icontains=query))
    if mercado:
        fornecedores = fornecedores.filter(mercado=mercado)

    contexto = {
        'fornecedores': fornecedores,
        'query': query,
        'mercado_filtro': mercado,
        'total_fornecedores': fornecedores.count(),
    }
    return render(request, 'core/lista_fornecedores.html', contexto)

@login_required
def adicionar_fornecedor(request):
    if request.method == 'POST':
        form = FornecedorForm(request.POST)
        if form.is_valid():
            fornecedor = form.save(commit=False)
            primeira_empresa = Empresa.objects.first()
            if primeira_empresa:
                fornecedor.empresa = primeira_empresa
            fornecedor.save()
            messages.success(request, f'Fornecedor "{fornecedor.nome}" cadastrado com sucesso!')
            return redirect('lista_fornecedores')
    else:
        form = FornecedorForm()

    return render(request, 'core/form_fornecedor.html', {'form': form, 'titulo': 'Adicionar Fornecedor'})

@login_required
def editar_fornecedor(request, pk):
    # SEGURANÇA: Garantir que o fornecedor pertence à empresa do usuário
    empresas = get_user_empresa(request.user)
    fornecedor = get_object_or_404(Fornecedor, pk=pk, empresa__in=empresas)

    if request.method == 'POST':
        form = FornecedorForm(request.POST, instance=fornecedor)
        if form.is_valid():
            form.save()
            messages.success(request, f'Fornecedor "{fornecedor.nome}" atualizado com sucesso!')
            return redirect('detalhe_fornecedor', pk=pk)
    else:
        form = FornecedorForm(instance=fornecedor)

    return render(request, 'core/form_fornecedor.html', {'form': form, 'titulo': f'Editar {fornecedor.nome}', 'fornecedor': fornecedor})

@login_required
def detalhe_fornecedor(request, pk):
    # SEGURANÇA: Garantir que o fornecedor pertence à empresa do usuário
    empresas = get_user_empresa(request.user)
    fornecedor = get_object_or_404(Fornecedor, pk=pk, empresa__in=empresas)
    return render(request, 'core/detalhe_fornecedor.html', {'fornecedor': fornecedor})

@login_required
def excluir_fornecedor(request, pk):
    # SEGURANÇA: Garantir que o fornecedor pertence à empresa do usuário
    empresas = get_user_empresa(request.user)
    fornecedor = get_object_or_404(Fornecedor, pk=pk, empresa__in=empresas)

    if request.method == 'POST':
        nome = fornecedor.nome
        fornecedor.delete()
        messages.success(request, f'Fornecedor "{nome}" excluído com sucesso!')
        return redirect('lista_fornecedores')
    return redirect('detalhe_fornecedor', pk=pk)

# === PWA MANIFEST DINÂMICO ===

def manifest_json(request):
    """Gera manifest.json dinâmico com URLs corretas para FORCE_SCRIPT_NAME"""
    from django.conf import settings
    from django.templatetags.static import static

    # Pega o prefixo base (FORCE_SCRIPT_NAME)
    base_url = settings.FORCE_SCRIPT_NAME if hasattr(settings, 'FORCE_SCRIPT_NAME') else ''

    manifest = {
        "name": "Blockline - Gestão Empresarial",
        "short_name": "Blockline",
        "description": "Sistema completo de gestão empresarial com controle de estoque, recebimento, expedição, kanban e ponto eletrônico",
        "start_url": f"{base_url}/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#4F46E5",
        "orientation": "portrait-primary",
        "scope": f"{base_url}/",
        "lang": "pt-BR",
        "dir": "ltr",
        "icons": [
            {
                "src": request.build_absolute_uri(static('icons/icon-72.png')),
                "sizes": "72x72",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": request.build_absolute_uri(static('icons/icon-96.png')),
                "sizes": "96x96",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": request.build_absolute_uri(static('icons/icon-128.png')),
                "sizes": "128x128",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": request.build_absolute_uri(static('icons/icon-144.png')),
                "sizes": "144x144",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": request.build_absolute_uri(static('icons/icon-152.png')),
                "sizes": "152x152",
                "type": "image/png",
                "purpose": "any"
            },
            {
                "src": request.build_absolute_uri(static('icons/icon-192.png')),
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable"
            },
            {
                "src": request.build_absolute_uri(static('icons/icon-512.png')),
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable"
            }
        ],
        "shortcuts": [
            {
                "name": "Bater Ponto",
                "short_name": "Ponto",
                "description": "Registrar ponto rapidamente",
                "url": f"{base_url}/ponto/",
                "icons": [{"src": request.build_absolute_uri(static('icons/icon-192.png')), "sizes": "192x192"}]
            },
            {
                "name": "Kanban",
                "short_name": "Kanban",
                "description": "Ver quadro Kanban",
                "url": f"{base_url}/kanban/",
                "icons": [{"src": request.build_absolute_uri(static('icons/icon-192.png')), "sizes": "192x192"}]
            },
            {
                "name": "Estoque",
                "short_name": "Estoque",
                "description": "Ver estoque",
                "url": f"{base_url}/estoque/",
                "icons": [{"src": request.build_absolute_uri(static('icons/icon-192.png')), "sizes": "192x192"}]
            }
        ],
        "prefer_related_applications": False
    }

    return JsonResponse(manifest, content_type='application/manifest+json')


# ==================================================
# VIEWS DE REQUISIÇÕES DE COMPRA
# ==================================================

@login_required
def lista_requisicoes(request):
    """Lista todas as requisições filtradas por status e nota fiscal"""
    from .models import RequisicaoCompra, Fornecedor, ProdutoFabricado
    from django.db.models import Q

    status_filtro = request.GET.get('status', 'todas')
    nf_filtro = request.GET.get('nf', '')

    if status_filtro == 'todas':
        requisicoes = RequisicaoCompra.objects.all()
    else:
        requisicoes = RequisicaoCompra.objects.filter(status=status_filtro)

    # Filtrar por nota fiscal se fornecido
    if nf_filtro:
        requisicoes = requisicoes.filter(nota_fiscal__icontains=nf_filtro)

    # Estatísticas
    total_pendente = RequisicaoCompra.objects.filter(status='pendente').count()
    total_aprovado = RequisicaoCompra.objects.filter(status='aprovado').count()
    total_comprado = RequisicaoCompra.objects.filter(status='comprado').count()
    total_recebido = RequisicaoCompra.objects.filter(status='recebido').count()
    total_rejeitado = RequisicaoCompra.objects.filter(status='rejeitado').count()

    fornecedores = Fornecedor.objects.all()
    produtos = ProdutoFabricado.objects.all().order_by('nome')

    context = {
        'requisicoes': requisicoes,
        'status_filtro': status_filtro,
        'nf_filtro': nf_filtro,
        'total_pendente': total_pendente,
        'total_aprovado': total_aprovado,
        'total_comprado': total_comprado,
        'total_recebido': total_recebido,
        'total_rejeitado': total_rejeitado,
        'fornecedores': fornecedores,
        'produtos': produtos,
    }

    return render(request, 'core/lista_requisicoes.html', context)


@login_required
@require_POST
def criar_requisicao(request):
    """Cria uma nova requisição de compra"""
    from .models import RequisicaoCompra

    produto_id = request.POST.get('produto')
    requisicao = RequisicaoCompra.objects.create(
        item=request.POST.get('item'),
        descricao=request.POST.get('descricao'),
        quantidade=request.POST.get('quantidade'),
        unidade=request.POST.get('unidade', 'un'),
        preco_estimado=request.POST.get('preco_estimado'),
        proposito=request.POST.get('proposito'),
        produto_id=produto_id if produto_id else None,
        link_item=request.POST.get('link_item', ''),
        requerente=request.user,
        status='pendente'
    )

    messages.success(request, f'Requisição "{requisicao.item}" criada com sucesso!')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def editar_requisicao(request, requisicao_id):
    """Edita uma requisição de compra - todos os campos"""
    from .models import RequisicaoCompra, HistoricoRequisicao
    from datetime import datetime
    from decimal import Decimal

    requisicao = get_object_or_404(RequisicaoCompra, id=requisicao_id)

    # Capturar valores antigos para comparação
    old_values = {
        'item': requisicao.item,
        'descricao': requisicao.descricao,
        'quantidade': requisicao.quantidade,
        'unidade': requisicao.unidade,
        'preco_estimado': requisicao.preco_estimado,
        'proposito': requisicao.proposito,
        'produto_id': requisicao.produto_id,
        'produto_nome': requisicao.produto.nome if requisicao.produto else None,
        'link_item': requisicao.link_item,
        'observacao_aprovacao': requisicao.observacao_aprovacao,
        'preco_real': requisicao.preco_real,
        'fornecedor_id': requisicao.fornecedor_id,
        'fornecedor_nome': requisicao.fornecedor.nome if requisicao.fornecedor else None,
        'nota_fiscal': requisicao.nota_fiscal,
        'data_entrega_prevista': requisicao.data_entrega_prevista,
        'observacao_recebimento': requisicao.observacao_recebimento,
        'forma_pagamento': requisicao.forma_pagamento,
        'quantidade_parcelas': requisicao.quantidade_parcelas,
        'dias_pagamento': requisicao.dias_pagamento,
        'dias_aviso_pagamento': requisicao.dias_aviso_pagamento,
    }

    changes = []

    # Atualizar campos básicos (sempre atualizar, registrar mudanças apenas se diferentes)
    new_item = request.POST.get('item')
    if new_item != old_values['item']:
        changes.append(f"Item alterado de '{old_values['item']}' para '{new_item}'")
    requisicao.item = new_item

    new_descricao = request.POST.get('descricao')
    if new_descricao != old_values['descricao']:
        changes.append(f"Descrição alterada")
    requisicao.descricao = new_descricao

    new_quantidade = int(request.POST.get('quantidade'))
    if new_quantidade != old_values['quantidade']:
        changes.append(f"Quantidade alterada de {old_values['quantidade']} para {new_quantidade}")
    requisicao.quantidade = new_quantidade

    new_unidade = request.POST.get('unidade', 'un')
    if new_unidade != old_values['unidade']:
        changes.append(f"Unidade alterada de '{old_values['unidade']}' para '{new_unidade}'")
    requisicao.unidade = new_unidade

    new_preco_estimado = Decimal(request.POST.get('preco_estimado'))
    if new_preco_estimado != old_values['preco_estimado']:
        changes.append(f"Preço estimado alterado de R$ {old_values['preco_estimado']} para R$ {new_preco_estimado}")
    requisicao.preco_estimado = new_preco_estimado

    new_proposito = request.POST.get('proposito')
    if new_proposito != old_values['proposito']:
        changes.append(f"Propósito alterado")
    requisicao.proposito = new_proposito

    new_produto_id = request.POST.get('produto')
    if new_produto_id:
        new_produto_id = int(new_produto_id)
    else:
        new_produto_id = None

    if new_produto_id != old_values['produto_id']:
        if new_produto_id:
            from .models import ProdutoFabricado
            produto = ProdutoFabricado.objects.get(id=new_produto_id)
            changes.append(f"Produto definido como '{produto.nome}'")
        else:
            changes.append(f"Produto removido")
    requisicao.produto_id = new_produto_id

    new_link_item = request.POST.get('link_item', '')
    if new_link_item != old_values['link_item']:
        if new_link_item:
            changes.append(f"Link do item alterado")
        elif old_values['link_item']:
            changes.append(f"Link do item removido")
    requisicao.link_item = new_link_item

    # Atualizar dados de pagamento
    forma_pagamento = request.POST.get('forma_pagamento', '')
    if forma_pagamento != old_values['forma_pagamento']:
        if forma_pagamento:
            changes.append(f"Forma de pagamento definida como '{dict(requisicao.FORMA_PAGAMENTO_CHOICES).get(forma_pagamento, forma_pagamento)}'")
        elif old_values['forma_pagamento']:
            changes.append(f"Forma de pagamento removida")
    requisicao.forma_pagamento = forma_pagamento

    if forma_pagamento == 'boleto':
        quantidade_parcelas = request.POST.get('quantidade_parcelas', '')
        if quantidade_parcelas:
            new_parcelas = int(quantidade_parcelas)
            if new_parcelas != old_values['quantidade_parcelas']:
                changes.append(f"Quantidade de parcelas alterada de {old_values['quantidade_parcelas'] or '0'} para {new_parcelas}")
            requisicao.quantidade_parcelas = new_parcelas

        dias_pagamento = request.POST.get('dias_pagamento', '')
        if dias_pagamento != old_values['dias_pagamento']:
            if dias_pagamento:
                changes.append(f"Dias de pagamento alterados")
            elif old_values['dias_pagamento']:
                changes.append(f"Dias de pagamento removidos")
        requisicao.dias_pagamento = dias_pagamento

        dias_aviso = request.POST.get('dias_aviso_pagamento', '')
        if dias_aviso:
            new_dias_aviso = int(dias_aviso)
            if new_dias_aviso != old_values['dias_aviso_pagamento']:
                changes.append(f"Dias de aviso alterados de {old_values['dias_aviso_pagamento'] or '3'} para {new_dias_aviso}")
            requisicao.dias_aviso_pagamento = new_dias_aviso

        # Upload de documento do boleto
        if request.FILES.get('documento_boleto'):
            changes.append(f"Documento do boleto atualizado")
            requisicao.documento_boleto = request.FILES['documento_boleto']

    # Upload de nota fiscal
    if request.FILES.get('documento_nota_fiscal'):
        changes.append(f"Documento da nota fiscal atualizado")
        requisicao.documento_nota_fiscal = request.FILES['documento_nota_fiscal']

    # Atualizar dados de aprovação
    observacao_aprovacao = request.POST.get('observacao_aprovacao', '')
    if observacao_aprovacao != old_values['observacao_aprovacao']:
        if observacao_aprovacao:
            changes.append(f"Observação de aprovação alterada")
        elif old_values['observacao_aprovacao']:
            changes.append(f"Observação de aprovação removida")
        requisicao.observacao_aprovacao = observacao_aprovacao

    # Atualizar documento de aprovação
    if request.FILES.get('documento_aprovacao'):
        changes.append(f"Documento de aprovação atualizado")
        requisicao.documento_aprovacao = request.FILES['documento_aprovacao']

    # Atualizar dados de compra
    preco_real = request.POST.get('preco_real', '')
    if preco_real:
        new_preco_real = Decimal(preco_real)
        if new_preco_real != old_values['preco_real']:
            changes.append(f"Preço real alterado de R$ {old_values['preco_real'] or '0.00'} para R$ {new_preco_real}")
        requisicao.preco_real = new_preco_real

    fornecedor_id = request.POST.get('fornecedor_id', '')
    if fornecedor_id:
        new_fornecedor_id = int(fornecedor_id)
        if new_fornecedor_id != old_values['fornecedor_id']:
            from .models import Fornecedor
            novo_fornecedor = Fornecedor.objects.get(id=new_fornecedor_id)
            changes.append(f"Fornecedor alterado de '{old_values['fornecedor_nome'] or 'Nenhum'}' para '{novo_fornecedor.nome}'")
        requisicao.fornecedor_id = new_fornecedor_id

    # Atualizar dados de recebimento
    nota_fiscal = request.POST.get('nota_fiscal', '')
    if nota_fiscal != old_values['nota_fiscal']:
        if nota_fiscal:
            changes.append(f"Nota fiscal alterada de '{old_values['nota_fiscal'] or 'Nenhuma'}' para '{nota_fiscal}'")
        elif old_values['nota_fiscal']:
            changes.append(f"Nota fiscal removida")
    requisicao.nota_fiscal = nota_fiscal

    data_entrega_str = request.POST.get('data_entrega_prevista', '')
    if data_entrega_str:
        try:
            new_data_entrega = datetime.strptime(data_entrega_str, '%Y-%m-%d').date()
            if new_data_entrega != old_values['data_entrega_prevista']:
                old_data_str = old_values['data_entrega_prevista'].strftime('%d/%m/%Y') if old_values['data_entrega_prevista'] else 'Nenhuma'
                changes.append(f"Data de entrega prevista alterada de {old_data_str} para {new_data_entrega.strftime('%d/%m/%Y')}")
            requisicao.data_entrega_prevista = new_data_entrega
        except ValueError:
            pass

    observacao_recebimento = request.POST.get('observacao_recebimento', '')
    if observacao_recebimento != old_values['observacao_recebimento']:
        if observacao_recebimento:
            changes.append(f"Observação de recebimento alterada")
        elif old_values['observacao_recebimento']:
            changes.append(f"Observação de recebimento removida")
    requisicao.observacao_recebimento = observacao_recebimento

    requisicao.save()

    # Criar registro no histórico se houver alterações
    if changes:
        HistoricoRequisicao.objects.create(
            requisicao=requisicao,
            usuario=request.user,
            tipo_alteracao="Edição Manual",
            descricao="\n".join(changes)
        )

    messages.success(request, f'Requisição "{requisicao.item}" atualizada com sucesso!')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def aprovar_requisicao(request, requisicao_id):
    """Aprova uma requisição"""
    from .models import RequisicaoCompra
    from django.utils import timezone

    requisicao = get_object_or_404(RequisicaoCompra, id=requisicao_id)

    if requisicao.status != 'pendente':
        messages.warning(request, 'Esta requisição já foi processada.')
        return redirect('lista_requisicoes')

    requisicao.status = 'aprovado'
    requisicao.aprovado_por = request.user
    requisicao.data_aprovacao = timezone.now()
    requisicao.observacao_aprovacao = request.POST.get('observacao', '')

    requisicao.save()

    messages.success(request, f'Requisição "{requisicao.item}" aprovada!')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def rejeitar_requisicao(request, requisicao_id):
    """Rejeita uma requisição"""
    from .models import RequisicaoCompra
    from django.utils import timezone

    requisicao = get_object_or_404(RequisicaoCompra, id=requisicao_id)

    if requisicao.status != 'pendente':
        messages.warning(request, 'Esta requisição já foi processada.')
        return redirect('lista_requisicoes')

    requisicao.status = 'rejeitado'
    requisicao.aprovado_por = request.user
    requisicao.data_aprovacao = timezone.now()
    requisicao.observacao_aprovacao = request.POST.get('observacao', 'Requisição rejeitada')
    requisicao.save()

    messages.warning(request, f'Requisição "{requisicao.item}" rejeitada.')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def rejeitar_compra(request, requisicao_id):
    """Rejeita uma compra aprovada, voltando status para rejeitado"""
    from .models import RequisicaoCompra
    from django.utils import timezone

    requisicao = get_object_or_404(RequisicaoCompra, id=requisicao_id)

    if requisicao.status != 'aprovado':
        messages.warning(request, 'Esta requisição não está aguardando compra.')
        return redirect('lista_requisicoes')

    observacao = request.POST.get('observacao', '')
    if not observacao:
        messages.error(request, 'É necessário informar o motivo da rejeição.')
        return redirect('lista_requisicoes')

    requisicao.status = 'rejeitado'
    requisicao.observacao_aprovacao = f"Compra rejeitada: {observacao}"
    requisicao.save()

    messages.warning(request, f'Compra da requisição "{requisicao.item}" rejeitada.')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def marcar_como_comprado(request, requisicao_id):
    """Marca requisição como comprada"""
    from .models import RequisicaoCompra
    from django.utils import timezone

    requisicao = get_object_or_404(RequisicaoCompra, id=requisicao_id)

    if requisicao.status != 'aprovado':
        messages.warning(request, 'Esta requisição precisa estar aprovada.')
        return redirect('lista_requisicoes')

    requisicao.status = 'comprado'
    requisicao.comprado_por = request.user
    requisicao.data_compra = timezone.now()
    requisicao.preco_real = request.POST.get('preco_real')

    # Processar fornecedor (cadastrado ou digitado)
    tipo_fornecedor = request.POST.get('tipo_fornecedor', 'cadastrado')
    if tipo_fornecedor == 'cadastrado':
        requisicao.fornecedor_id = request.POST.get('fornecedor_id')
        requisicao.fornecedor_nome_digitado = None
    else:
        requisicao.fornecedor_id = None
        requisicao.fornecedor_nome_digitado = request.POST.get('fornecedor_nome_digitado', '')

    requisicao.nota_fiscal = request.POST.get('nota_fiscal', '')

    # Processar data de entrega prevista
    data_entrega_str = request.POST.get('data_entrega_prevista')
    if data_entrega_str:
        from datetime import datetime
        try:
            requisicao.data_entrega_prevista = datetime.strptime(data_entrega_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Dados de pagamento
    requisicao.forma_pagamento = request.POST.get('forma_pagamento', '')

    # Se for boleto, capturar campos específicos
    if requisicao.forma_pagamento == 'boleto':
        requisicao.quantidade_parcelas = request.POST.get('quantidade_parcelas') or None
        dias_aviso = request.POST.get('dias_aviso_pagamento')
        requisicao.dias_aviso_pagamento = int(dias_aviso) if dias_aviso else 3

        # Processar tipo de dias de pagamento
        tipo_dias = request.POST.get('tipo_dias_pagamento', '')
        requisicao.tipo_dias_pagamento = tipo_dias

        if tipo_dias == '15_em_15' or tipo_dias == '30_em_30':
            # Calcular datas baseadas na data inicial
            data_inicial_str = request.POST.get('data_inicial_boleto', '')
            quantidade_str = request.POST.get('quantidade_parcelas_calc', '3')

            if data_inicial_str:
                from datetime import datetime, timedelta
                try:
                    data_inicial = datetime.strptime(data_inicial_str, '%Y-%m-%d').date()
                    quantidade = int(quantidade_str) if quantidade_str else 3
                    intervalo_dias = 15 if tipo_dias == '15_em_15' else 30

                    # Calcular as datas somando o intervalo
                    datas_list = []
                    for i in range(quantidade):
                        data_vencimento = data_inicial + timedelta(days=intervalo_dias * i)
                        datas_list.append(data_vencimento.strftime('%Y-%m-%d'))

                    requisicao.dias_pagamento = ', '.join(datas_list)
                    requisicao.quantidade_parcelas = quantidade
                except ValueError:
                    # Se houver erro, deixar em branco
                    requisicao.dias_pagamento = ''
            else:
                requisicao.dias_pagamento = ''

        elif tipo_dias == 'especificos':
            # Usar datas selecionadas pelo usuário
            requisicao.dias_pagamento = request.POST.get('dias_pagamento', '')

        # Upload de documento do boleto
        if request.FILES.get('documento_boleto'):
            requisicao.documento_boleto = request.FILES['documento_boleto']

    # Upload de nota fiscal (para todas as formas de pagamento)
    if request.FILES.get('documento_nota_fiscal'):
        requisicao.documento_nota_fiscal = request.FILES['documento_nota_fiscal']

    requisicao.save()

    # Se forma de pagamento é boleto, criar parcelas automaticamente
    if requisicao.forma_pagamento == 'boleto' and requisicao.dias_pagamento:
        from .models import ParcelaBoleto
        from datetime import datetime

        # Parsear datas de vencimento
        dias_str = requisicao.dias_pagamento.strip()
        datas_vencimento = []

        for parte in dias_str.split(','):
            parte = parte.strip()
            # Tentar como data completa (YYYY-MM-DD)
            if '-' in parte and len(parte) >= 8:
                try:
                    data_vencimento = datetime.strptime(parte, '%Y-%m-%d').date()
                    datas_vencimento.append(data_vencimento)
                except ValueError:
                    pass

        if datas_vencimento:
            # Calcular valor da parcela
            valor_total = requisicao.preco_real or requisicao.preco_estimado or Decimal('0')
            quantidade = len(datas_vencimento)
            valor_parcela = valor_total / quantidade if quantidade > 0 else Decimal('0')

            # Criar parcelas
            for i, data_vencimento in enumerate(datas_vencimento, start=1):
                ParcelaBoleto.objects.create(
                    requisicao=requisicao,
                    numero_parcela=i,
                    data_vencimento=data_vencimento,
                    valor=valor_parcela,
                    pago=False
                )

    messages.success(request, f'Requisição "{requisicao.item}" marcada como comprada!')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def marcar_como_recebido(request, requisicao_id):
    """Marca requisição como recebida"""
    from .models import RequisicaoCompra
    from django.utils import timezone

    requisicao = get_object_or_404(RequisicaoCompra, id=requisicao_id)

    if requisicao.status != 'comprado':
        messages.warning(request, 'Esta requisição precisa estar marcada como comprada.')
        return redirect('lista_requisicoes')

    requisicao.status = 'recebido'
    requisicao.recebido_por = request.user
    requisicao.data_recebimento = timezone.now()
    requisicao.observacao_recebimento = request.POST.get('observacao', '')

    # Processar nota fiscal
    nota_fiscal = request.POST.get('nota_fiscal', '')
    if nota_fiscal:
        requisicao.nota_fiscal = nota_fiscal

    # Processar data de entrega prevista
    data_entrega_str = request.POST.get('data_entrega_prevista')
    if data_entrega_str:
        from datetime import datetime
        try:
            requisicao.data_entrega_prevista = datetime.strptime(data_entrega_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    requisicao.save()

    messages.success(request, f'Requisição "{requisicao.item}" marcada como recebida!')
    return redirect('lista_requisicoes')


@login_required
@require_POST
def marcar_parcela_paga(request, parcela_id):
    """Marca parcela de boleto como paga"""
    from .models import ParcelaBoleto

    # Verificar se usuário é do financeiro
    if not hasattr(request.user, 'perfil') or not request.user.perfil.is_financeiro:
        messages.error(request, 'Apenas usuários do financeiro podem marcar parcelas como pagas.')
        return redirect('lista_requisicoes')

    parcela = get_object_or_404(ParcelaBoleto, id=parcela_id)

    if parcela.pago:
        messages.warning(request, 'Esta parcela já foi marcada como paga.')
        return redirect('lista_requisicoes')

    # Marcar como pago
    parcela.pago = True
    parcela.pago_por = request.user

    # Data do pagamento
    data_pagamento_str = request.POST.get('data_pagamento')
    if data_pagamento_str:
        from datetime import datetime
        parcela.data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
    else:
        parcela.data_pagamento = timezone.now().date()

    # Comprovante
    if 'comprovante' in request.FILES:
        parcela.comprovante = request.FILES['comprovante']

    # Observações
    parcela.observacoes = request.POST.get('observacoes', '')

    parcela.save()

    messages.success(request, f'Parcela {parcela.numero_parcela} marcada como paga com sucesso!')
    return redirect('lista_requisicoes')


@login_required
def alertas_boletos_api(request):
    """API que retorna alertas de boletos para usuários do financeiro"""
    from .models import ParcelaBoleto
    from datetime import timedelta
    from django.http import JsonResponse

    # Apenas para usuários do financeiro
    if not hasattr(request.user, 'perfil') or not request.user.perfil.is_financeiro:
        return JsonResponse({'alertas': [], 'total': 0})

    hoje = timezone.now().date()

    # Buscar parcelas não pagas com vencimento próximo (próximos 7 dias)
    parcelas_alerta = ParcelaBoleto.objects.filter(
        pago=False,
        data_vencimento__gte=hoje,
        data_vencimento__lte=hoje + timedelta(days=7)
    ).select_related('requisicao').order_by('data_vencimento')

    alertas = []
    for parcela in parcelas_alerta:
        dias_restantes = (parcela.data_vencimento - hoje).days
        alertas.append({
            'id': parcela.id,
            'requisicao_id': parcela.requisicao.id,
            'item': parcela.requisicao.item,
            'parcela': f"{parcela.numero_parcela}/{parcela.requisicao.quantidade_parcelas or '?'}",
            'valor': float(parcela.valor),
            'data_vencimento': parcela.data_vencimento.strftime('%d/%m/%Y'),
            'dias_restantes': dias_restantes,
            'urgente': dias_restantes <= 3
        })

    return JsonResponse({'alertas': alertas, 'total': len(alertas)})


# ==================================================
# VIEWS DE GASTOS DE VIAGEM
# ==================================================

@login_required
def lista_gastos_viagem(request):
    """Lista todos os gastos de viagem"""
    from .models import GastoViagem
    from django.db.models import Sum
    from decimal import Decimal

    gastos = GastoViagem.objects.all().select_related('usuario')

    # Calcular total de gastos
    total_gastos = gastos.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    # Calcular totais separados por status de envio ao financeiro
    total_enviado = gastos.filter(enviado_financeiro=True).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    total_pendente = gastos.filter(enviado_financeiro=False).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    context = {
        'gastos': gastos,
        'total_gastos': total_gastos,
        'total_enviado': total_enviado,
        'total_pendente': total_pendente,
    }

    return render(request, 'core/lista_gastos_viagem.html', context)


@login_required
@require_POST
def criar_gasto_viagem(request):
    """Cria um novo gasto de viagem"""
    from .models import GastoViagem
    from datetime import datetime

    gasto = GastoViagem.objects.create(
        usuario=request.user,
        valor=request.POST.get('valor'),
        descricao=request.POST.get('descricao'),
        destino=request.POST.get('destino', ''),
        categoria=request.POST.get('categoria', ''),
        nota_fiscal=request.POST.get('nota_fiscal', ''),
    )

    # Processar imagem
    if request.FILES.get('imagem'):
        gasto.imagem = request.FILES['imagem']

    # Processar data da viagem
    data_viagem_str = request.POST.get('data_viagem')
    if data_viagem_str:
        try:
            gasto.data_viagem = datetime.strptime(data_viagem_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    gasto.save()

    messages.success(request, 'Gasto de viagem registrado com sucesso!')
    return redirect('lista_gastos_viagem')


@login_required
@require_POST
def excluir_gasto_viagem(request, gasto_id):
    """Exclui um gasto de viagem"""
    from .models import GastoViagem

    gasto = get_object_or_404(GastoViagem, id=gasto_id)
    gasto.delete()

    messages.success(request, 'Gasto de viagem excluído com sucesso!')
    return redirect('lista_gastos_viagem')


@login_required
@require_POST
def toggle_enviado_financeiro_viagem(request, gasto_id):
    """Alterna o status de envio ao financeiro para gastos de viagem"""
    from .models import GastoViagem

    gasto = get_object_or_404(GastoViagem, id=gasto_id)
    gasto.enviado_financeiro = not gasto.enviado_financeiro
    gasto.save()

    status = "enviado ao" if gasto.enviado_financeiro else "marcado como não enviado ao"
    messages.success(request, f'Gasto {status} financeiro!')
    return redirect('lista_gastos_viagem')


@login_required
@require_POST
def editar_gasto_viagem(request, gasto_id):
    """Edita um gasto de viagem"""
    from .models import GastoViagem
    from datetime import datetime

    gasto = get_object_or_404(GastoViagem, id=gasto_id)

    gasto.valor = request.POST.get('valor')
    gasto.descricao = request.POST.get('descricao')
    gasto.destino = request.POST.get('destino', '')
    gasto.categoria = request.POST.get('categoria', '')
    gasto.nota_fiscal = request.POST.get('nota_fiscal', '')

    # Processar imagem se foi enviada
    if request.FILES.get('imagem'):
        gasto.imagem = request.FILES['imagem']

    # Processar data da viagem
    data_viagem_str = request.POST.get('data_viagem')
    if data_viagem_str:
        try:
            gasto.data_viagem = datetime.strptime(data_viagem_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    gasto.save()

    messages.success(request, 'Gasto de viagem atualizado com sucesso!')
    return redirect('lista_gastos_viagem')


@login_required
@require_POST
def exportar_gastos_viagem_excel(request):
    """Exporta gastos de viagem selecionados para Excel com formatação"""
    from .models import GastoViagem
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from decimal import Decimal

    # Obter IDs selecionados
    gasto_ids = request.POST.get('gasto_ids', '').split(',')
    gasto_ids = [int(id) for id in gasto_ids if id]

    # Buscar gastos
    gastos = GastoViagem.objects.filter(id__in=gasto_ids).order_by('data_gasto')

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos de Viagem"

    # Estilos aprimorados
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=13)

    # Bordas
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Cores alternadas
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Fonte padrão
    data_font = Font(name='Calibri', size=11)

    # Cabeçalhos
    headers = ["Data Registro", "Nota Fiscal", "Categoria", "Destino", "Data Viagem", "Descrição", "Valor (R$)", "Link Imagem"]
    ws.append(headers)

    # Formatar cabeçalho
    ws.row_dimensions[1].height = 30
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20  # Data Registro
    ws.column_dimensions['B'].width = 18  # Nota Fiscal
    ws.column_dimensions['C'].width = 22  # Categoria
    ws.column_dimensions['D'].width = 25  # Destino
    ws.column_dimensions['E'].width = 16  # Data Viagem
    ws.column_dimensions['F'].width = 60  # Descrição
    ws.column_dimensions['G'].width = 16  # Valor
    ws.column_dimensions['H'].width = 50  # Link

    # Adicionar dados
    total = Decimal('0.00')
    row_num = 2

    for idx, gasto in enumerate(gastos):
        fill = light_fill if idx % 2 == 0 else white_fill

        # Altura dinâmica
        descricao_lines = len(gasto.descricao) // 80 + 1
        row_height = max(35, min(descricao_lines * 15, 100))
        ws.row_dimensions[row_num].height = row_height

        # Dados
        data = [
            gasto.data_gasto.strftime('%d/%m/%Y %H:%M'),
            gasto.nota_fiscal or '-',
            gasto.categoria or '-',
            gasto.destino or '-',
            gasto.data_viagem.strftime('%d/%m/%Y') if gasto.data_viagem else '-',
            gasto.descricao,
            float(gasto.valor),
            ''
        ]

        ws.append(data)

        # Formatar células
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.border = border
            cell.font = data_font

            if col_num in [1, 2, 3, 4, 5]:  # Data, NF, Categoria, Destino, Data Viagem
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 6:  # Descrição
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_num == 7:  # Valor
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 8:  # Link
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adicionar link para imagem
        if gasto.imagem:
            try:
                img_url = request.build_absolute_uri(gasto.imagem.url)
                img_cell = ws.cell(row=row_num, column=8)
                img_cell.value = img_url
                img_cell.hyperlink = img_url
                img_cell.font = Font(name='Calibri', size=10, color="0563C1", underline="single")
                img_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            except Exception:
                img_cell = ws.cell(row=row_num, column=8)
                img_cell.value = "-"
        else:
            img_cell = ws.cell(row=row_num, column=8)
            img_cell.value = "-"
            img_cell.font = Font(name='Calibri', size=11, color="999999")

        total += gasto.valor
        row_num += 1

    # Linha de total
    row_num += 1
    ws.row_dimensions[row_num].height = 35

    ws.merge_cells(f'A{row_num}:F{row_num}')
    total_label_cell = ws.cell(row=row_num, column=1)
    total_label_cell.value = "TOTAL GERAL"
    total_label_cell.font = Font(name='Calibri', bold=True, size=14, color="000000")
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_label_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    total_label_cell.border = border

    total_value_cell = ws.cell(row=row_num, column=7)
    total_value_cell.value = float(total)
    total_value_cell.number_format = 'R$ #,##0.00'
    total_value_cell.font = Font(name='Calibri', bold=True, size=14, color="000000")
    total_value_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_value_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    total_value_cell.border = border

    for col in [2, 3, 4, 5, 6, 8]:
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.border = border

    # Recursos avançados
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{row_num-1}'
    ws.sheet_view.zoomScale = 90

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=gastos_viagem.xlsx'

    wb.save(response)
    return response


# ==================================================
# VIEWS DE GASTOS DE CAIXA INTERNO
# ==================================================

@login_required
def lista_gastos_caixa(request):
    """Lista todos os gastos do caixa interno"""
    from .models import GastoCaixaInterno
    from django.db.models import Sum
    from decimal import Decimal

    gastos = GastoCaixaInterno.objects.all().select_related('usuario')

    # Calcular total de gastos
    total_gastos = gastos.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    # Calcular totais separados por status de envio ao financeiro
    total_enviado = gastos.filter(enviado_financeiro=True).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    total_pendente = gastos.filter(enviado_financeiro=False).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    context = {
        'gastos': gastos,
        'total_gastos': total_gastos,
        'total_enviado': total_enviado,
        'total_pendente': total_pendente,
    }

    return render(request, 'core/lista_gastos_caixa.html', context)


@login_required
@require_POST
def criar_gasto_caixa(request):
    """Cria um novo gasto do caixa interno"""
    from .models import GastoCaixaInterno

    gasto = GastoCaixaInterno.objects.create(
        usuario=request.user,
        valor=request.POST.get('valor'),
        descricao=request.POST.get('descricao'),
        categoria=request.POST.get('categoria', ''),
        nota_fiscal=request.POST.get('nota_fiscal', ''),
    )

    # Processar imagem
    if request.FILES.get('imagem'):
        gasto.imagem = request.FILES['imagem']

    gasto.save()

    messages.success(request, 'Gasto de caixa registrado com sucesso!')
    return redirect('lista_gastos_caixa')


@login_required
@require_POST
def excluir_gasto_caixa(request, gasto_id):
    """Exclui um gasto do caixa interno"""
    from .models import GastoCaixaInterno

    gasto = get_object_or_404(GastoCaixaInterno, id=gasto_id)
    gasto.delete()

    messages.success(request, 'Gasto de caixa excluído com sucesso!')
    return redirect('lista_gastos_caixa')


@login_required
@require_POST
def toggle_enviado_financeiro(request, gasto_id):
    """Alterna o status de envio ao financeiro"""
    from .models import GastoCaixaInterno

    gasto = get_object_or_404(GastoCaixaInterno, id=gasto_id)
    gasto.enviado_financeiro = not gasto.enviado_financeiro
    gasto.save()

    status = "enviado ao" if gasto.enviado_financeiro else "marcado como não enviado ao"
    messages.success(request, f'Gasto {status} financeiro!')
    return redirect('lista_gastos_caixa')


@login_required
@require_POST
def editar_gasto_caixa(request, gasto_id):
    """Edita um gasto do caixa interno"""
    from .models import GastoCaixaInterno

    gasto = get_object_or_404(GastoCaixaInterno, id=gasto_id)

    gasto.valor = request.POST.get('valor')
    gasto.descricao = request.POST.get('descricao')
    gasto.categoria = request.POST.get('categoria', '')
    gasto.nota_fiscal = request.POST.get('nota_fiscal', '')

    # Processar imagem se foi enviada
    if request.FILES.get('imagem'):
        gasto.imagem = request.FILES['imagem']

    gasto.save()

    messages.success(request, 'Gasto de caixa atualizado com sucesso!')
    return redirect('lista_gastos_caixa')


@login_required
@require_POST
def exportar_gastos_excel(request):
    """Exporta gastos selecionados para Excel com formatação"""
    from .models import GastoCaixaInterno
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from decimal import Decimal

    # Obter IDs selecionados
    gasto_ids = request.POST.get('gasto_ids', '').split(',')
    gasto_ids = [int(id) for id in gasto_ids if id]

    # Buscar gastos
    gastos = GastoCaixaInterno.objects.filter(id__in=gasto_ids).order_by('data_gasto')

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos de Caixa"

    # Estilos aprimorados
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Azul mais escuro
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=13)

    # Bordas mais visíveis
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Cores alternadas mais suaves
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Cinza muito claro
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Fonte padrão para dados
    data_font = Font(name='Calibri', size=11)

    # Cabeçalhos
    headers = ["Data", "Nota Fiscal", "Categoria", "Descrição", "Valor (R$)", "Link Imagem"]
    ws.append(headers)

    # Formatar cabeçalho com altura otimizada
    ws.row_dimensions[1].height = 30
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Ajustar largura das colunas para melhor leitura
    ws.column_dimensions['A'].width = 20  # Data - mais espaço
    ws.column_dimensions['B'].width = 18  # Nota Fiscal - mais legível
    ws.column_dimensions['C'].width = 22  # Categoria - confortável
    ws.column_dimensions['D'].width = 60  # Descrição - maior para ler melhor
    ws.column_dimensions['E'].width = 16  # Valor - espaço adequado
    ws.column_dimensions['F'].width = 50  # URL da imagem - espaço para ver o link completo

    # Adicionar dados
    total = Decimal('0.00')
    row_num = 2

    for idx, gasto in enumerate(gastos):
        # Determinar cor da linha (alternada)
        fill = light_fill if idx % 2 == 0 else white_fill

        # Altura da linha dinâmica baseada no tamanho da descrição
        # Aproximadamente 15 pixels por linha de texto
        descricao_lines = len(gasto.descricao) // 80 + 1  # Estimar linhas
        row_height = max(35, min(descricao_lines * 15, 100))  # Entre 35 e 100
        ws.row_dimensions[row_num].height = row_height

        # Dados
        data = [
            gasto.data_gasto.strftime('%d/%m/%Y %H:%M'),
            gasto.nota_fiscal or '-',
            gasto.categoria or '-',
            gasto.descricao,
            float(gasto.valor),
            ''  # Coluna para link da imagem
        ]

        ws.append(data)

        # Formatar células com estilos aprimorados
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.border = border
            cell.font = data_font

            # Alinhamentos específicos por coluna
            if col_num == 1:  # Data
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [2, 3]:  # Nota Fiscal e Categoria
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 4:  # Descrição
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_num == 5:  # Valor
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num == 6:  # Link
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adicionar link para imagem se existir
        if gasto.imagem:
            try:
                # Obter URL completa da imagem
                img_url = request.build_absolute_uri(gasto.imagem.url)

                # Adicionar URL completa visível (funciona em qualquer visualizador)
                img_cell = ws.cell(row=row_num, column=6)
                img_cell.value = img_url
                img_cell.hyperlink = img_url
                img_cell.font = Font(name='Calibri', size=10, color="0563C1", underline="single")
                img_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            except Exception as e:
                # Se houver erro, indicar sem imagem
                img_cell = ws.cell(row=row_num, column=6)
                img_cell.value = "-"
                img_cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Sem imagem
            img_cell = ws.cell(row=row_num, column=6)
            img_cell.value = "-"
            img_cell.font = Font(name='Calibri', size=11, color="999999")
            img_cell.alignment = Alignment(horizontal="center", vertical="center")

        total += gasto.valor
        row_num += 1

    # Adicionar linha de total aprimorada
    row_num += 1
    ws.row_dimensions[row_num].height = 35

    # Mesclar células para o rótulo
    ws.merge_cells(f'A{row_num}:D{row_num}')
    total_label_cell = ws.cell(row=row_num, column=1)
    total_label_cell.value = "TOTAL GERAL"
    total_label_cell.font = Font(name='Calibri', bold=True, size=14, color="000000")
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_label_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Amarelo suave
    total_label_cell.border = border

    # Célula do valor total
    total_value_cell = ws.cell(row=row_num, column=5)
    total_value_cell.value = float(total)
    total_value_cell.number_format = 'R$ #,##0.00'
    total_value_cell.font = Font(name='Calibri', bold=True, size=14, color="000000")
    total_value_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_value_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    total_value_cell.border = border

    # Aplicar formatação nas células vazias da linha de total
    for col in [2, 3, 4, 6]:
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        cell.border = border

    # Congelar painéis (fixar cabeçalho)
    ws.freeze_panes = 'A2'

    # Habilitar auto-filtro no cabeçalho
    ws.auto_filter.ref = f'A1:F{row_num-1}'

    # Configurar zoom padrão (90% para melhor visualização)
    ws.sheet_view.zoomScale = 90

    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=gastos_caixa.xlsx'

    wb.save(response)
    return response


# ==================================================
# DASHBOARD DE COMPRAS
# ==================================================

@login_required
def dashboard_compras(request):
    """
    Dashboard consolidado de gastos de compras com:
    - Totais mensais por categoria (viagem, caixa, compras)
    - Filtro por produto
    - Filtro por mês/ano
    - Gráfico de evolução dos últimos 12 meses
    """
    from .models import GastoViagem, GastoCaixaInterno, RequisicaoCompra, ProdutoFabricado
    from django.db.models import Sum, F, ExpressionWrapper, DecimalField
    from datetime import datetime, timedelta
    from decimal import Decimal
    import calendar

    # 1. Obter filtros (produto, mês, ano)
    produto_id = request.GET.get('produto_id', None)

    hoje = datetime.now()

    # Obter mês e ano dos parâmetros GET (ou usar mês/ano atual como padrão)
    try:
        mes_selecionado = int(request.GET.get('mes', hoje.month))
    except (ValueError, TypeError):
        mes_selecionado = hoje.month

    try:
        ano_selecionado = int(request.GET.get('ano', hoje.year))
    except (ValueError, TypeError):
        ano_selecionado = hoje.year

    # Validar mês (1-12)
    if mes_selecionado < 1 or mes_selecionado > 12:
        mes_selecionado = hoje.month

    # 2. Query base para requisições (já filtra por data_compra não nula)
    requisicoes_qs = RequisicaoCompra.objects.filter(
        data_compra__isnull=False,
        preco_real__isnull=False
    )

    # Aplicar filtro de produto se selecionado
    if produto_id:
        requisicoes_qs = requisicoes_qs.filter(produto_id=produto_id)

    # 3. Calcular totais do mês selecionado
    mes_atual = mes_selecionado
    ano_atual = ano_selecionado

    # Total de Gastos de Viagem (mês atual)
    total_viagem_mes = GastoViagem.objects.filter(
        data_gasto__month=mes_atual,
        data_gasto__year=ano_atual
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    # Total de Gastos de Caixa (mês atual)
    total_caixa_mes = GastoCaixaInterno.objects.filter(
        data_gasto__month=mes_atual,
        data_gasto__year=ano_atual
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    # Total de Compras (mês atual) - preco_real × quantidade
    compras_mes = requisicoes_qs.filter(
        data_compra__month=mes_atual,
        data_compra__year=ano_atual
    ).annotate(
        valor_total=ExpressionWrapper(
            F('preco_real') * F('quantidade'),
            output_field=DecimalField()
        )
    ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0.00')

    # Total Geral do Mês
    total_mes = total_viagem_mes + total_caixa_mes + compras_mes

    # 5. Preparar dados para gráfico (últimos 12 meses)
    meses_labels = []
    dados_viagem = []
    dados_caixa = []
    dados_compras = []

    for i in range(11, -1, -1):
        # Calcular mês e ano (subtraindo meses do atual)
        meses_atras = i
        data_ref = hoje.replace(day=1) - timedelta(days=meses_atras * 30)
        mes = data_ref.month
        ano = data_ref.year

        # Label do mês
        try:
            nome_mes = calendar.month_abbr[mes]  # Jan, Feb, etc
        except:
            nome_mes = str(mes)
        meses_labels.append(f"{nome_mes}/{str(ano)[2:]}")

        # Viagem
        viagem = GastoViagem.objects.filter(
            data_gasto__month=mes,
            data_gasto__year=ano
        ).aggregate(total=Sum('valor'))['total'] or 0
        dados_viagem.append(float(viagem))

        # Caixa
        caixa = GastoCaixaInterno.objects.filter(
            data_gasto__month=mes,
            data_gasto__year=ano
        ).aggregate(total=Sum('valor'))['total'] or 0
        dados_caixa.append(float(caixa))

        # Compras
        compras = requisicoes_qs.filter(
            data_compra__month=mes,
            data_compra__year=ano
        ).annotate(
            valor_total=ExpressionWrapper(
                F('preco_real') * F('quantidade'),
                output_field=DecimalField()
            )
        ).aggregate(total=Sum('valor_total'))['total'] or 0
        dados_compras.append(float(compras))

    # 6. Lista de produtos para filtro
    produtos = ProdutoFabricado.objects.all().order_by('nome')

    # 7. Preparar lista de anos disponíveis (últimos 5 anos + ano atual + próximo ano)
    ano_inicial = hoje.year - 5
    ano_final = hoje.year + 1
    anos_disponiveis = list(range(ano_inicial, ano_final + 1))

    # 8. Nome do mês selecionado para exibição
    try:
        nome_mes_selecionado = calendar.month_name[mes_selecionado]  # Janeiro, Fevereiro, etc.
    except:
        nome_mes_selecionado = f"Mês {mes_selecionado}"

    context = {
        'total_viagem_mes': total_viagem_mes,
        'total_caixa_mes': total_caixa_mes,
        'total_compras_mes': compras_mes,
        'total_mes': total_mes,
        'meses_labels': meses_labels,
        'dados_viagem': dados_viagem,
        'dados_caixa': dados_caixa,
        'dados_compras': dados_compras,
        'produtos': produtos,
        'produto_selecionado': produto_id,
        'mes_selecionado': mes_selecionado,
        'ano_selecionado': ano_selecionado,
        'anos_disponiveis': anos_disponiveis,
        'nome_mes_selecionado': nome_mes_selecionado,
    }

    return render(request, 'core/dashboard_compras.html', context)


# ==================================================
# VIEWS DE PERFIL DO USUÁRIO
# ==================================================

@login_required
def perfil_usuario(request):
    """Exibe o perfil do usuário logado"""
    context = {
        'user': request.user,
    }
    return render(request, 'core/perfil_usuario.html', context)


@login_required
def alterar_senha(request):
    """Permite ao usuário alterar sua senha"""
    if request.method == 'POST':
        senha_atual = request.POST.get('senha_atual')
        nova_senha = request.POST.get('nova_senha')
        confirmar_senha = request.POST.get('confirmar_senha')

        # Validar senha atual
        if not request.user.check_password(senha_atual):
            messages.error(request, 'Senha atual incorreta!')
            return redirect('perfil_usuario')

        # Validar nova senha
        if len(nova_senha) < 6:
            messages.error(request, 'A nova senha deve ter pelo menos 6 caracteres!')
            return redirect('perfil_usuario')

        # Validar confirmação
        if nova_senha != confirmar_senha:
            messages.error(request, 'As senhas não coincidem!')
            return redirect('perfil_usuario')

        # Alterar senha
        request.user.set_password(nova_senha)
        request.user.save()

        # Manter usuário logado após mudança de senha
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)

        messages.success(request, 'Senha alterada com sucesso!')
        return redirect('perfil_usuario')

    return redirect('perfil_usuario')


@login_required
def editar_perfil(request):
    """Permite ao usuário editar informações básicas do perfil"""
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()

        # Validar email
        if email and User.objects.exclude(id=request.user.id).filter(email=email).exists():
            messages.error(request, 'Este email já está sendo usado por outro usuário!')
            return redirect('perfil_usuario')

        # Atualizar dados
        request.user.first_name = first_name
        request.user.last_name = last_name
        request.user.email = email
        request.user.save()

        messages.success(request, 'Perfil atualizado com sucesso!')
        return redirect('perfil_usuario')

    return redirect('perfil_usuario')


# ==================================================
# VIEWS - SISTEMA DE NOTIFICAÇÕES
# ==================================================

@login_required
def listar_notificacoes(request):
    """Listar notificações do usuário (AJAX)"""
    from .models import Notificacao
    from datetime import datetime, timedelta

    def tempo_relativo(data_criacao):
        """Calcula tempo relativo em português"""
        agora = timezone.now()
        diff = agora - data_criacao

        if diff < timedelta(minutes=1):
            return "Agora"
        elif diff < timedelta(hours=1):
            minutos = int(diff.total_seconds() / 60)
            return f"{minutos} min atrás"
        elif diff < timedelta(days=1):
            horas = int(diff.total_seconds() / 3600)
            return f"{horas}h atrás"
        elif diff < timedelta(days=7):
            dias = diff.days
            return f"{dias} dia{'s' if dias > 1 else ''} atrás"
        elif diff < timedelta(days=30):
            semanas = diff.days // 7
            return f"{semanas} semana{'s' if semanas > 1 else ''} atrás"
        else:
            return data_criacao.strftime('%d/%m/%Y')

    notificacoes = Notificacao.objects.filter(usuario=request.user).order_by('-criado_em')[:20]

    data = {
        'notificacoes': [
            {
                'id': n.id,
                'tipo': n.tipo,
                'titulo': n.titulo,
                'mensagem': n.mensagem,
                'lida': n.lida,
                'criado_em': n.criado_em.strftime('%d/%m/%Y %H:%M'),
                'tempo_relativo': tempo_relativo(n.criado_em),
                'task_id': n.task.id if n.task else None,
            }
            for n in notificacoes
        ],
        'nao_lidas': Notificacao.objects.filter(usuario=request.user, lida=False).count()
    }

    return JsonResponse(data)


@login_required
@require_POST
def marcar_notificacao_lida(request, notificacao_id):
    """Marcar notificação como lida (AJAX)"""
    from .models import Notificacao

    try:
        notificacao = Notificacao.objects.get(id=notificacao_id, usuario=request.user)
        notificacao.marcar_como_lida()

        return JsonResponse({
            'success': True,
            'nao_lidas': Notificacao.objects.filter(usuario=request.user, lida=False).count()
        })
    except Notificacao.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notificação não encontrada'}, status=404)


@login_required
@require_POST
def marcar_todas_lidas(request):
    """Marcar todas notificações como lidas (AJAX)"""
    from .models import Notificacao

    Notificacao.objects.filter(usuario=request.user, lida=False).update(
        lida=True,
        lida_em=timezone.now()
    )

    return JsonResponse({
        'success': True,
        'nao_lidas': 0
    })


@login_required
def contar_notificacoes(request):
    """Contar notificações não lidas (AJAX)"""
    from .models import Notificacao

    count = Notificacao.objects.filter(usuario=request.user, lida=False).count()

    return JsonResponse({'nao_lidas': count})